import hashlib
import io
import unicodedata
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import Articulo, CargaInicialInventario, Kardex


ENCABEZADOS = {
    'codigo': 'codigo',
    'renglon': 'renglon',
    'descripcion': 'nombre',
    'articulo': 'nombre',
    'nombre': 'nombre',
    'categoria': 'categoria',
    'unidad': 'unidad',
    'unidad de medida': 'unidad',
    'cantidad': 'cantidad',
    'cantidad inicial': 'cantidad',
    'existencia': 'cantidad',
    'costo unitario': 'costo_unitario',
    'precio unitario': 'costo_unitario',
    'total': 'total_excel',
}


def normalizar_encabezado(valor):
    texto = unicodedata.normalize('NFKD', str(valor or ''))
    texto = ''.join(c for c in texto if not unicodedata.combining(c)).lower().strip()
    return ' '.join(texto.replace('_', ' ').split())


def _decimal(valor, predeterminado=None):
    if valor in (None, ''):
        return predeterminado
    texto = str(valor).strip().replace('Q', '').replace(' ', '')
    if ',' in texto and '.' in texto:
        texto = texto.replace(',', '')
    else:
        texto = texto.replace(',', '.')
    return Decimal(texto)


def analizar_excel(contenido):
    """Lee y valida el libro sin persistir ningún dato."""
    libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))
    indice_encabezado = None
    columnas = {}
    for indice, fila in enumerate(filas[:20]):
        candidatas = {pos: ENCABEZADOS.get(normalizar_encabezado(valor)) for pos, valor in enumerate(fila)}
        candidatas = {pos: nombre for pos, nombre in candidatas.items() if nombre}
        if {'codigo', 'nombre', 'cantidad'}.issubset(candidatas.values()):
            indice_encabezado, columnas = indice, candidatas
            break
    if indice_encabezado is None:
        raise ValueError('No se encontraron las columnas obligatorias Código, Artículo y Cantidad.')

    resultado, codigos = [], set()
    for numero_fila, valores in enumerate(filas[indice_encabezado + 1:], indice_encabezado + 2):
        if not any(valor not in (None, '') for valor in valores):
            continue
        dato = {nombre: valores[pos] if pos < len(valores) else None for pos, nombre in columnas.items()}
        codigo = str(dato.get('codigo') or '').strip()
        nombre = str(dato.get('nombre') or '').strip()
        errores, advertencias = [], []
        if not codigo:
            errores.append('código vacío')
        elif codigo.casefold() in codigos:
            errores.append('código duplicado en el archivo')
        else:
            codigos.add(codigo.casefold())
        if not nombre:
            errores.append('artículo/descripción vacío')
        try:
            cantidad_decimal = _decimal(dato.get('cantidad'))
            if cantidad_decimal is None or cantidad_decimal < 0 or cantidad_decimal != cantidad_decimal.to_integral_value():
                raise InvalidOperation
            cantidad = int(cantidad_decimal)
        except (InvalidOperation, ValueError, TypeError):
            cantidad = 0
            errores.append('cantidad inválida')
        try:
            costo = _decimal(dato.get('costo_unitario'), Decimal('0'))
            if costo < 0:
                raise InvalidOperation
            costo = costo.quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError, TypeError):
            costo = Decimal('0.00')
            errores.append('costo unitario inválido')
        total = Decimal(cantidad) * costo
        try:
            total_excel = _decimal(dato.get('total_excel'))
            if total_excel is not None and total_excel.quantize(Decimal('0.01')) != total:
                advertencias.append('El total del Excel no coincide con cantidad x costo unitario.')
        except (InvalidOperation, ValueError, TypeError):
            advertencias.append('El total indicado en el Excel no es válido.')

        articulo = Articulo.objects.filter(codigo__iexact=codigo).first() if codigo else None
        tiene_movimientos = bool(articulo and (articulo.detalles_factura.exists() or Kardex.objects.filter(articulo=articulo).exists()))
        if tiene_movimientos:
            advertencias.append('El artículo ya tiene movimientos registrados. Verifique antes de actualizar.')
        estado = 'error' if errores else ('actualizar' if articulo else 'nuevo')
        resultado.append({
            'fila': numero_fila, 'codigo': codigo, 'nombre': nombre,
            'renglon': str(dato.get('renglon') or '').strip(),
            'categoria': str(dato.get('categoria') or '').strip(),
            'unidad': str(dato.get('unidad') or '').strip() or 'Unidad',
            'cantidad': cantidad, 'costo_unitario': str(costo), 'total': str(total),
            'estado': estado, 'errores': errores, 'advertencias': advertencias,
            'tiene_movimientos': tiene_movimientos,
        })
    return resultado


def hash_archivo(contenido):
    return hashlib.sha256(contenido).hexdigest()


def carga_ya_procesada(contenido):
    return CargaInicialInventario.objects.filter(hash_archivo=hash_archivo(contenido)).exists()
