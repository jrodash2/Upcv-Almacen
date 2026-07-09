from datetime import datetime, timezone
from venv import logger
from django.forms import IntegerField
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import Group, User
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .form import DependenciaForm, DetalleFacturaForm, DetalleRequerimientoForm, DetalleRequerimientoFormSet, Form1hForm, PerfilForm, ProgramaForm, RequerimientoForm, UserCreateForm, UserEditForm, UserCreateForm, UbicacionForm, UnidadDeMedidaForm, CategoriaForm, ProveedorForm, ArticuloForm, DepartamentoForm, SerieForm, AsignacionDetalleFacturaForm, UsuarioDepartamentoForm, InstitucionForm, SolicitudRequerimientoForm, DetalleSolicitudRequerimientoFormSet, DivisionAlmacenForm, DivisionUbicacionForm, DivisionArticuloForm
from .models import existencia_general_articulo, ContadorDetalleFactura, DetalleFactura, DetalleRequerimiento, HistorialTransferencia, InventarioDetalle, LineaLibre, Perfil, Requerimiento, Ubicacion, UnidadDeMedida, Categoria, Proveedor, Articulo, Departamento, Kardex, AsignacionDetalleFactura, Movimiento, FraseMotivacional, Serie, form1h, Dependencia, Programa, LineaReservada, UsuarioDepartamento, Institucion, SolicitudRequerimiento, DetalleSolicitudRequerimiento, DivisionAlmacen, DivisionUbicacion, DivisionArticulo, DivisionArticuloUbicacion
from django.views.generic import CreateView
from django.views.generic import ListView
from django.urls import reverse_lazy
from django.http import Http404, HttpResponseNotAllowed, JsonResponse
from django.core.exceptions import ValidationError
from django.contrib import messages
from .utils import reservar_lineas
from .models import LineaLibre, ContadorDetalleFactura, LineaReservada
from django.views.decorators.http import require_POST
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .models import Serie
from django.db import models
from django.db.models import Sum, F, Value, Count, Q, Case, When, OuterRef, Subquery, IntegerField
from django.contrib.auth.decorators import login_required, user_passes_test
from collections import defaultdict
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
import json
from django.contrib.auth.models import Group
from .utils import grupo_requerido
from django.views.decorators.http import require_GET
from django.db.models.functions import Coalesce
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import render
from .models import DetalleFactura, AsignacionDetalleFactura, Articulo
from django.template.loader import render_to_string
from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from weasyprint import HTML
from django.db.models.functions import Cast, TruncWeek
from django.utils import timezone
from datetime import timedelta
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import datetime
from django.core.mail import send_mail
from django.conf import settings
from django.utils.html import strip_tags
from decimal import Decimal
from datetime import datetime  
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import re

from django.core.mail import BadHeaderError
from smtplib import SMTPException

@login_required
@grupo_requerido('Administrador', 'Almacen')
def exportar_detalle_factura_excel(request, form1h_id):
    form1h_instance = get_object_or_404(form1h, id=form1h_id)
    detalles = DetalleFactura.objects.filter(form1h=form1h_instance)

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Factura"

    # Información superior
    ws["A1"] = "Dependencia:"
    ws["B1"] = str(form1h_instance.dependencia)
    ws["D1"] = "Factura:"
    ws["E1"] = form1h_instance.numero_factura

    ws["A2"] = "Programa:"
    ws["B2"] = str(form1h_instance.programa)
    ws["D2"] = "Fecha:"
    ws["E2"] = form1h_instance.fecha_factura.strftime("%d/%m/%Y")

    ws["A3"] = "Proveedor:"
    ws["B3"] = str(form1h_instance.proveedor)
    ws["D3"] = "Orden Compra:"
    ws["E3"] = form1h_instance.orden_compra

    # Encabezados
    headers = [
        "Cantidad", "Detalle", "Renglón", "Folio Almacén",
        "Precio Unidad", "Valor Total", "Folio Inventario", "Nomenclatura"
    ]
    ws.append([])
    ws.append(headers)

    # Filas de datos
    for d in detalles:
        inventarios = d.inventarios.all()
        if inventarios.exists():
            for i, inv in enumerate(inventarios):
                if i == 0:
                    ws.append([
                        d.cantidad,
                        d.articulo.nombre,
                        d.renglon,
                        getattr(d, 'numero_linea', ''),
                        f"Q{d.precio_unitario}",
                        f"Q{d.precio_total}",
                        inv.folio_inventario,
                        inv.nomenclatura
                    ])
                else:
                    ws.append([
                        "", "", "", "", "", "",
                        inv.folio_inventario,
                        inv.nomenclatura
                    ])
        else:
            ws.append([
                d.cantidad,
                d.articulo.nombre,
                d.renglon,
                getattr(d, 'numero_linea', ''),
                f"Q{d.precio_unitario}",
                f"Q{d.precio_total}",
                "-", "-"
            ])

    # Alinear texto en celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Ajustar ancho de columnas automáticamente (opcional)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Nombre archivo usando los campos correctos
    serie_texto = str(form1h_instance.serie.serie) if form1h_instance.serie else ''
    numero = str(form1h_instance.numero_serie) if form1h_instance.numero_serie else ''
    nombre_archivo = f"formulario_1h_{serie_texto} {numero}".strip()

    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response



def enviar_correo_asignacion(articulo, departamento, usuario):
    asunto = f"Asignación de artículo: {articulo.nombre}"
    mensaje = f"Se ha asignado el artículo {articulo.nombre} a {departamento.nombre}. para ver mas detalles ingrese con sus credenciales a la Aplicación de Gestión de Almacen"
    destinatario = usuario.email

    try:
        send_mail(
            asunto,
            mensaje,
            'informatica@upcv.gob.gt',
            [destinatario],
            fail_silently=False,
        )
        logger.info(f"Correo enviado correctamente a {destinatario}")
    except BadHeaderError:
        logger.error("Header del correo es incorrecto")
    except SMTPException as e:
        logger.error(f"Error al enviar el correo: {e}")


@login_required
@grupo_requerido('Administrador', 'Almacen')
def exportar_detalle_factura_pdf(request, form1h_id):
    form1h_instance = get_object_or_404(form1h, id=form1h_id)
    detalles = DetalleFactura.objects.filter(form1h=form1h_instance)
    institucion = Institucion.objects.first()  # Opcional, si ya lo usas en otros PDF

    total_factura = form1h_instance.calcular_total_factura()

    html_string = render_to_string('almacen/pdf_detalle_factura.html', {
        'form1h_instance': form1h_instance,
        'detalles': detalles,
        'institucion': institucion,
        'total_factura': total_factura,
    })

    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="detalle_factura_{form1h_instance.id}.pdf"'
    return response

def libro_ingresos_pdf(request):
    # Filtros para fecha (si es que se aplican)
    fecha_inicio = request.GET.get('fecha_inicio', '2025-01-01')
    fecha_fin = request.GET.get('fecha_fin', '2025-12-31')

    # Convertir las fechas usando strptime (aquí corregimos la importación)
    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

    # Filtrar facturas dentro del rango de fechas
    facturas = form1h.objects.filter(fecha_ingreso__range=(fecha_inicio, fecha_fin), estado='confirmado')

    # Consultar todos los detalles de factura
    detalles_factura = DetalleFactura.objects.filter(form1h__in=facturas).select_related('articulo')

    # Agrupar detalles de factura por artículo
    ingresos_por_articulo = detalles_factura.values('articulo__nombre').annotate(
        total_ingresos=Sum('precio_total'), total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    # Crear PDF con orientación horizontal y tamaño oficio
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="libro_ingresos.pdf"'

    # Tamaño de página Oficio (8.5 x 14 pulgadas) en orientación horizontal
    p = canvas.Canvas(response, pagesize=(14*inch, 8.5*inch))  # 14x8.5 pulgadas

    width, height = 14*inch, 8.5*inch  # Dimensiones de la página

    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1*inch, height - 1*inch, "Libro de Ingresos")

    p.setFont("Helvetica", 10)
    p.drawString(1*inch, height - 1.3*inch, f"Fecha inicio: {fecha_inicio} - Fecha fin: {fecha_fin}")

    # Tabla encabezados
    p.drawString(1*inch, height - 1.8*inch, "Artículo")
    p.drawString(4*inch, height - 1.8*inch, "Cantidad Total")
    p.drawString(7*inch, height - 1.8*inch, "Monto Total")

    # Dibujar los datos
    y = height - 2*inch
    for ingreso in ingresos_por_articulo:
        p.drawString(1*inch, y, ingreso['articulo__nombre'])
        p.drawString(4*inch, y, str(ingreso['total_cantidad']))
        p.drawString(7*inch, y, f"{ingreso['total_ingresos']:.2f}")
        y -= 0.3*inch  # Ajustar la distancia entre filas
        if y < 1*inch:
            p.showPage()
            y = height - 1*inch

    p.showPage()
    p.save()

    return response








def historial_transferencias(request):
    # Obtén todos los registros del historial de transferencias
    transferencias = HistorialTransferencia.objects.all().order_by('-fecha_transferencia')
    
    return render(request, 'almacen/historial_transferencias.html', {'transferencias': transferencias})

def obtener_stock_disponible_por_departamento(departamento):
    asignaciones = (
        AsignacionDetalleFactura.objects
        .filter(destino=departamento)
        .values('articulo_id')
        .annotate(total_asignado=Coalesce(Sum('cantidad_asignada'), 0))
    )
    despachados = (
        DetalleRequerimiento.objects
        .filter(requerimiento__departamento=departamento)
        .values('articulo_id')
        .annotate(total_despachado=Coalesce(Sum('cantidad_despachada'), 0))
    )
    despachados_dict = {d['articulo_id']: d['total_despachado'] for d in despachados}

    stock_disponible = {}
    for a in asignaciones:
        disponible = a['total_asignado'] - despachados_dict.get(a['articulo_id'], 0)
        if disponible > 0:
            stock_disponible[a['articulo_id']] = disponible

    asignaciones_division = (
        DivisionArticuloUbicacion.objects
        .filter(ubicacion=departamento, activo=True, division_articulo__activo=True)
        .select_related('division_articulo__articulo')
    )
    for asignacion in asignaciones_division:
        disponible = asignacion.disponible_ubicacion
        if disponible > 0:
            articulo_id = asignacion.division_articulo.articulo_id
            stock_disponible[articulo_id] = stock_disponible.get(articulo_id, 0) + disponible

    return stock_disponible

@login_required
@grupo_requerido('Gestor', 'Departamento', 'Administrador', 'Almacen')
def articulos_asignados(request, departamento_id):
    try:
        departamento = Departamento.objects.get(
            id=departamento_id,
            usuariodepartamento__usuario=request.user
        )
        stock_disponible = obtener_stock_disponible_por_departamento(departamento)
        articulos = Articulo.objects.filter(id__in=stock_disponible.keys())
        tipos_solicitud = request.GET.getlist('tipos_solicitud') or request.GET.getlist('tipos_solicitud[]')
        if tipos_solicitud:
            categoria_query = Q()
            filtros_categoria = {
                'bienes': 'bien',
                'suministros': 'suministro',
                'insumos': 'insumo',
            }
            for tipo in tipos_solicitud:
                categoria_query |= Q(categoria__nombre__icontains=filtros_categoria.get(tipo, tipo))
            articulos = articulos.filter(categoria_query)
        articulos = articulos.order_by('nombre')

        asignaciones_division = {
            item.division_articulo.articulo_id: item
            for item in DivisionArticuloUbicacion.objects.filter(
                ubicacion=departamento,
                activo=True,
                division_articulo__activo=True,
            ).select_related('division_articulo__division', 'division_articulo__articulo')
            if item.disponible_ubicacion > 0
        }
        data = {
            'articulos': [
                {
                    'id': articulo.id,
                    'codigo': articulo.codigo,
                    'nombre': articulo.nombre,
                    'categoria': articulo.categoria.nombre if articulo.categoria else 'S/C',
                    'renglon_presupuestario': articulo.renglon_presupuestario or 'S/R',
                    'division': asignaciones_division.get(articulo.id).division_articulo.division.nombre if asignaciones_division.get(articulo.id) else '',
                    'cantidad_disponible': str(stock_disponible.get(articulo.id, 0)),
                    'cantidad_asignada': str(asignaciones_division.get(articulo.id).cantidad_asignada if asignaciones_division.get(articulo.id) else stock_disponible.get(articulo.id, 0)),
                    'cantidad_reservada': str(asignaciones_division.get(articulo.id).cantidad_reservada if asignaciones_division.get(articulo.id) else 0),
                    'cantidad_consumida': str(asignaciones_division.get(articulo.id).cantidad_consumida if asignaciones_division.get(articulo.id) else 0),
                }
                for articulo in articulos
            ]
        }
        return JsonResponse(data)
    except Departamento.DoesNotExist:
        return JsonResponse({'error': 'Departamento no encontrado'}, status=404)
    
    
@login_required
@grupo_requerido('Administrador')
def editar_institucion(request):
    institucion = Institucion.objects.first()  # Solo debería haber una

    if request.method == 'POST':
        form = InstitucionForm(request.POST, request.FILES, instance=institucion)
        if form.is_valid():
            form.save()
            messages.success(request, "Datos institucionales actualizados correctamente.")
            return redirect('almacen:editar_institucion')  # Reemplaza con la URL real
    else:
        form = InstitucionForm(instance=institucion)

    return render(request, 'almacen/editar_institucion.html', {'form': form})


from django.db.models import F

@login_required
# @grupo_requerido('Administrador', 'Almacen') # Descomenta si usas este decorador
@transaction.atomic
def despachar_requerimiento(request, requerimiento_id):
    if not request.user.groups.filter(name__in=['Administrador', 'Almacen']).exists():
        messages.error(request, "No tienes permiso para despachar requerimientos.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento_id)

    requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)

    if requerimiento.estado != 'enviado':
        messages.warning(request, "Solo se pueden despachar requerimientos en estado 'enviado'.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

    if request.method == 'POST':
        cantidades_dict = {}
        for key in request.POST:
            if key.startswith('cantidades[') and key.endswith(']'):
                detalle_id = key[11:-1]
                try:
                    cantidades_dict[int(detalle_id)] = int(request.POST[key])
                except ValueError:
                    cantidades_dict[int(detalle_id)] = 0

        detalles = DetalleRequerimiento.objects.filter(requerimiento=requerimiento)

        for detalle in detalles:
            cantidad_a_despachar = cantidades_dict.get(detalle.id, 0)
            
            # Validación: no se puede despachar más de lo solicitado
            if cantidad_a_despachar > detalle.cantidad:
                messages.error(request, f"No puedes despachar más de lo solicitado para {detalle.articulo.nombre}.")
                return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)
            
            # Validación: no despachar cantidades negativas
            if cantidad_a_despachar < 0:
                messages.error(request, f"La cantidad a despachar para {detalle.articulo.nombre} no puede ser negativa.")
                return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

            # Despacho desde los lotes disponibles
            articulo = detalle.articulo
            cantidad_restante = cantidad_a_despachar

            lotes = DetalleFactura.objects.filter(
                articulo=articulo,
                form1h__estado='confirmado'
            ).annotate(
                total_salidas=Coalesce(
                    Sum('kardex__cantidad', filter=Q(kardex__tipo_movimiento='SALIDA', kardex__fuente_factura__isnull=False)), 0
                ),
                stock_disponible=F('cantidad') - F('total_salidas')
            ).filter(
                stock_disponible__gt=0
            ).order_by('fecha_vencimiento', 'id')  # FIFO

            for lote in lotes:
                if cantidad_restante <= 0:
                    break

                cantidad_despacho = min(cantidad_restante, lote.stock_disponible)
                cantidad_restante -= cantidad_despacho

            if cantidad_restante > 0:
                messages.error(request, f"No hay suficiente stock disponible para despachar {detalle.articulo.nombre}.")
                return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

            # Actualizar el detalle. Esto dispara la señal para crear el Kardex
            detalle.cantidad_despachada = cantidad_a_despachar
            detalle.estado = 'despachado' if cantidad_a_despachar == detalle.cantidad else 'parcial'
            detalle.save()
            solicitud_origen = SolicitudRequerimiento.objects.filter(requerimiento=requerimiento).first()
            if solicitud_origen:
                detalle_solicitud = solicitud_origen.detalles.filter(articulo=detalle.articulo, division_articulo_ubicacion__isnull=False).first()
                if detalle_solicitud:
                    asignacion = detalle_solicitud.division_articulo_ubicacion
                    mover = min(Decimal(cantidad_a_despachar), asignacion.cantidad_reservada)
                    asignacion.cantidad_reservada -= mover
                    asignacion.cantidad_consumida += mover
                    asignacion.save(update_fields=['cantidad_reservada', 'cantidad_consumida', 'fecha_actualizacion'])

        # Verificar si todos los detalles fueron despachados completamente
        if all(d.estado == 'despachado' for d in detalles):
            requerimiento.estado = 'despachado'
            requerimiento.fecha_despachado = timezone.now()
            requerimiento.despachado_por = request.user
        else:
            requerimiento.estado = 'parcial'
        requerimiento.save()

        messages.success(request, "Requerimiento despachado exitosamente.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

    messages.error(request, "Método no permitido.")
    return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)
    
from django.db.models import Q

@login_required
def crear_requerimiento(request):
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    es_almacen = request.user.groups.filter(name='Almacen').exists()
    es_departamento = request.user.groups.filter(name='Departamento').exists()
    es_gestor = request.user.groups.filter(name='Gestor').exists()

    if not (es_admin or es_almacen or es_departamento or es_gestor):
        messages.error(request, "No tiene permiso para consultar requerimientos.")
        return redirect('almacen:acceso_denegado')

    puede_crear_requerimiento = es_admin or es_departamento

    if request.method == "POST" and not puede_crear_requerimiento:
        messages.error(request, "No tiene permiso para crear requerimientos directos.")
        return redirect('almacen:seguimiento_requerimientos' if es_gestor else 'almacen:crear_requerimiento')

    form = RequerimientoForm(request.POST or None, usuario=request.user)

    if es_admin or es_almacen:
        requerimientos = Requerimiento.objects.filter(
            estado__in=['enviado', 'despachado', 'rechazado', 'parcial', 'pendiente']
        ).order_by('-fecha_creacion')
    else:
        departamentos_usuario = UsuarioDepartamento.objects.filter(usuario=request.user).values_list('departamento', flat=True)
        filtros = Q(departamento__in=departamentos_usuario)
        if es_departamento:
            filtros |= Q(creado_por=request.user)
        if es_gestor:
            filtros |= Q(solicitudes_origen__usuario_solicitante=request.user)

        requerimientos = Requerimiento.objects.filter(filtros).distinct().order_by('-fecha_creacion')

    resumen_requerimientos = {
        'total': requerimientos.count(),
        'pendientes': requerimientos.filter(estado='pendiente').count(),
        'enviados': requerimientos.filter(estado='enviado').count(),
        'despachados': requerimientos.filter(estado='despachado').count(),
        'rechazados': requerimientos.filter(estado='rechazado').count(),
    }

    if request.method == "POST":
        if form.is_valid():
            nuevo = form.save(commit=False)
            nuevo.creado_por = request.user
            if not nuevo.departamento_id:
                asignacion = UsuarioDepartamento.objects.filter(usuario=request.user).first()
                if asignacion:
                    nuevo.departamento = asignacion.departamento
            nuevo.save()
            messages.success(request, "Requerimiento creado exitosamente.")
            return redirect('almacen:detalle_requerimiento', requerimiento_id=nuevo.id)
        messages.error(request, "Por favor revisa el formulario.")

    return render(request, 'almacen/crear_requerimiento.html', {
        'form': form,
        'requerimientos': requerimientos,
        'mostrar_modal': request.method == "POST" and not form.is_valid(),
        'es_admin': es_admin,
        'es_almacen': es_almacen,
        'es_departamento': es_departamento,
        'es_gestor': es_gestor,
        'puede_crear_requerimiento': puede_crear_requerimiento,
        'resumen_requerimientos': resumen_requerimientos,
    })


@login_required
@grupo_requerido('Gestor')
def crear_solicitud_requerimiento(request):
    form = SolicitudRequerimientoForm(request.POST or None, usuario=request.user)
    departamentos_qs = form.fields['departamento'].queryset
    departamento = departamentos_qs.first() if request.method == 'GET' else None
    if request.method == 'POST':
        departamento_id = request.POST.get('departamento')
        if departamento_id:
            departamento = departamentos_qs.filter(id=departamento_id).first()

    stock_disponible = obtener_stock_disponible_por_departamento(departamento) if departamento else {}
    tipos_solicitud = request.POST.getlist('tipos_solicitud') if request.method == 'POST' else []

    formset = DetalleSolicitudRequerimientoFormSet(
        request.POST or None,
        queryset=DetalleSolicitudRequerimiento.objects.none(),
        prefix='detalles',
        departamento=departamento,
        stock_disponible=stock_disponible,
        tipos_solicitud=tipos_solicitud
    )

    if request.method == 'POST' and not tipos_solicitud:
        messages.error(request, "Debe seleccionar al menos un tipo de solicitud.")

    if request.method == 'POST' and not stock_disponible:
        messages.error(request, "No tiene artículos asignados disponibles para solicitar. Comuníquese con el administrador o con su departamento.")

    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        try:
            with transaction.atomic():
                solicitud = form.save(commit=False)
                solicitud.usuario_solicitante = request.user
                solicitud.save()
                for f in formset:
                    if f.cleaned_data and not f.cleaned_data.get('DELETE'):
                        detalle = f.save(commit=False)
                        detalle.solicitud = solicitud
                        asignaciones_division = (
                            DivisionArticuloUbicacion.objects
                            .select_for_update()
                            .filter(
                                ubicacion=solicitud.departamento,
                                division_articulo__articulo=detalle.articulo,
                                activo=True,
                                division_articulo__activo=True,
                            )
                            .select_related('division_articulo')
                            .order_by('division_articulo__division__nombre', 'id')
                        )
                        cantidad_detalle = Decimal(detalle.cantidad)
                        asignacion_division = next((asig for asig in asignaciones_division if asig.disponible_ubicacion >= cantidad_detalle), None)
                        if asignaciones_division.exists() and not asignacion_division:
                            raise ValidationError(f"La cantidad solicitada supera el disponible de división para {detalle.articulo.codigo}.")
                        if asignacion_division:
                            asignacion_division.cantidad_reservada += cantidad_detalle
                            asignacion_division.save(update_fields=['cantidad_reservada', 'fecha_actualizacion'])
                            detalle.division_articulo_ubicacion = asignacion_division
                        detalle.save()
        except ValidationError as e:
            messages.error(request, e.message_dict if hasattr(e, 'message_dict') else e.messages)
        else:
            messages.success(request, "Solicitud creada correctamente.")
            return redirect('almacen:listado_solicitudes_gestor')
    return render(request, 'almacen/crear_solicitud_requerimiento.html', {
        'form': form,
        'formset': formset,
        'stock_disponible': {str(k): str(v) for k, v in stock_disponible.items()},
        'stock_dict': {str(k): str(v) for k, v in stock_disponible.items()},
        'sin_stock_disponible': not bool(stock_disponible),
        'selected_tipos': tipos_solicitud,
    })


@login_required
@grupo_requerido('Gestor')
def listado_solicitudes_gestor(request):
    solicitudes = SolicitudRequerimiento.objects.filter(usuario_solicitante=request.user).order_by('-creado_en')
    return render(request, 'almacen/listado_solicitudes_gestor.html', {'solicitudes': solicitudes})


@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento')
def bandeja_solicitudes_requerimiento(request):
    solicitudes = SolicitudRequerimiento.objects.all().order_by('-creado_en')
    if request.user.groups.filter(name='Departamento').exists():
        deptos = UsuarioDepartamento.objects.filter(usuario=request.user).values_list('departamento_id', flat=True)
        solicitudes = solicitudes.filter(departamento_id__in=deptos)
    return render(request, 'almacen/bandeja_solicitudes_requerimiento.html', {'solicitudes': solicitudes})


@login_required
def detalle_solicitud_requerimiento(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudRequerimiento, id=solicitud_id)
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    es_almacen = request.user.groups.filter(name='Almacen').exists()
    es_departamento = request.user.groups.filter(name='Departamento').exists()
    es_gestor = request.user.groups.filter(name='Gestor').exists()
    tiene_departamento_asignado = UsuarioDepartamento.objects.filter(
        usuario=request.user,
        departamento=solicitud.departamento
    ).exists()

    if not (es_admin or es_almacen or (es_departamento and tiene_departamento_asignado) or (es_gestor and solicitud.usuario_solicitante == request.user)):
        messages.error(request, "No tiene permiso para consultar esta solicitud.")
        return redirect('almacen:acceso_denegado')

    return render(request, 'almacen/detalle_solicitud_requerimiento.html', {
        'solicitud': solicitud,
        'puede_procesar_solicitud': es_admin or es_almacen or (es_departamento and tiene_departamento_asignado),
    })


def usuario_puede_ver_solicitud_requerimiento(user, solicitud):
    es_admin = user.is_superuser or user.groups.filter(name='Administrador').exists()
    es_almacen = user.groups.filter(name='Almacen').exists()
    es_departamento = user.groups.filter(name='Departamento').exists()
    es_gestor = user.groups.filter(name='Gestor').exists()
    tiene_departamento_asignado = UsuarioDepartamento.objects.filter(
        usuario=user,
        departamento=solicitud.departamento
    ).exists()

    return (
        es_admin
        or es_almacen
        or (es_departamento and tiene_departamento_asignado)
        or (es_gestor and solicitud.usuario_solicitante_id == user.id)
    )


@login_required
def solicitud_requerimiento_pdf(request, solicitud_id):
    solicitud = get_object_or_404(
        SolicitudRequerimiento.objects.select_related('departamento', 'usuario_solicitante'),
        id=solicitud_id
    )

    if not usuario_puede_ver_solicitud_requerimiento(request.user, solicitud):
        messages.error(request, "No tiene permiso para imprimir esta solicitud.")
        return redirect('almacen:listado_solicitudes_gestor')

    detalles = list(solicitud.detalles.select_related('articulo', 'articulo__unidad_medida').all())
    min_filas = 10
    filas_vacias = range(max(0, min_filas - len(detalles)))

    html_string = render_to_string(
        'almacen/solicitudes/solicitud_requerimiento_pdf.html',
        {
            'solicitud': solicitud,
            'detalles': detalles,
            'filas_vacias': filas_vacias,
            'cargo_solicitante': '',
            'institucion': Institucion.objects.first(),
        },
        request=request
    )

    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="solicitud_{solicitud.id}.pdf"'
    return response


@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento')
@require_POST
def convertir_solicitud_en_requerimiento(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudRequerimiento.objects.select_related('requerimiento', 'departamento'), id=solicitud_id)
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    es_almacen = request.user.groups.filter(name='Almacen').exists()
    tiene_departamento_asignado = UsuarioDepartamento.objects.filter(usuario=request.user, departamento=solicitud.departamento).exists()
    if not (es_admin or es_almacen or tiene_departamento_asignado):
        messages.error(request, "No tiene permiso para convertir esta solicitud.")
        return redirect('almacen:acceso_denegado')
    if solicitud.requerimiento_id:
        messages.warning(request, "Esta solicitud ya fue convertida.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=solicitud.requerimiento_id)
    if solicitud.estado != 'pendiente':
        messages.warning(request, "La solicitud ya fue procesada.")
        return redirect('almacen:detalle_solicitud_requerimiento', solicitud_id=solicitud.id)
    try:
        with transaction.atomic():
            detalles = list(solicitud.detalles.select_related('articulo', 'division_articulo_ubicacion').all())
            if not detalles:
                raise ValidationError("La solicitud no tiene detalles para convertir.")
            requerimiento = Requerimiento.objects.create(
                departamento=solicitud.departamento,
                motivo=solicitud.justificacion or solicitud.observaciones,
                creado_por=solicitud.usuario_solicitante,
                estado='pendiente'
            )
            for detalle in detalles:
                if not detalle.division_articulo_ubicacion_id:
                    disponible = AsignacionDetalleFactura.objects.filter(
                        destino=solicitud.departamento, articulo=detalle.articulo
                    ).aggregate(total=Coalesce(Sum('cantidad_asignada'), 0))['total'] or 0
                    despachado = DetalleRequerimiento.objects.filter(
                        requerimiento__departamento=solicitud.departamento, articulo=detalle.articulo
                    ).aggregate(total=Coalesce(Sum('cantidad_despachada'), 0))['total'] or 0
                    if detalle.cantidad > (disponible - despachado):
                        raise ValidationError(f"No hay stock suficiente para {detalle.articulo.codigo}.")
                DetalleRequerimiento.objects.create(
                    requerimiento=requerimiento,
                    articulo=detalle.articulo,
                    cantidad=detalle.cantidad,
                    observacion=detalle.observacion
                )
            solicitud.estado = 'convertida'
            solicitud.requerimiento = requerimiento
            solicitud.convertido_por = request.user
            solicitud.convertido_en = timezone.now()
            solicitud.save(update_fields=['estado', 'requerimiento', 'convertido_por', 'convertido_en', 'actualizado_en'])
    except ValidationError as e:
        messages.error(request, e.message_dict if hasattr(e, 'message_dict') else e.messages)
        return redirect('almacen:detalle_solicitud_requerimiento', solicitud_id=solicitud.id)
    messages.success(request, "Solicitud convertida en requerimiento correctamente.")
    return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)


@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento')
def rechazar_solicitud_requerimiento(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudRequerimiento, id=solicitud_id)
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    es_almacen = request.user.groups.filter(name='Almacen').exists()
    tiene_departamento_asignado = UsuarioDepartamento.objects.filter(usuario=request.user, departamento=solicitud.departamento).exists()
    if not (es_admin or es_almacen or tiene_departamento_asignado):
        messages.error(request, "No tiene permiso para rechazar esta solicitud.")
        return redirect('almacen:acceso_denegado')
    if solicitud.estado != 'pendiente':
        messages.warning(request, "La solicitud ya fue procesada.")
        return redirect('almacen:bandeja_solicitudes_requerimiento')
    motivo = request.POST.get('motivo_rechazo', '').strip()
    if request.method == 'POST' and motivo:
        solicitud.estado = 'rechazada'
        solicitud.motivo_rechazo = motivo
        solicitud.rechazado_por = request.user
        solicitud.rechazado_en = timezone.now()
        solicitud.save()
        for detalle in solicitud.detalles.select_related('division_articulo_ubicacion'):
            if detalle.division_articulo_ubicacion_id:
                asignacion = detalle.division_articulo_ubicacion
                asignacion.cantidad_reservada = max(Decimal('0'), asignacion.cantidad_reservada - Decimal(detalle.cantidad))
                asignacion.save(update_fields=['cantidad_reservada', 'fecha_actualizacion'])
        messages.success(request, "Solicitud rechazada correctamente.")
        return redirect('almacen:bandeja_solicitudes_requerimiento')
    return render(request, 'almacen/rechazar_solicitud_requerimiento.html', {'solicitud': solicitud})

@login_required
def enviar_requerimiento(request, requerimiento_id):
    if not request.user.groups.filter(name__in=['Administrador', 'Almacen', 'Departamento']).exists():
        messages.error(request, "No tiene permiso para realizar esta acción.")
        return redirect('almacen:acceso_denegado')

    requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)

    if requerimiento.estado != 'confirmado':  # Evita cambiar si ya fue confirmado
        requerimiento.estado = 'enviado'
        requerimiento.save()
        messages.success(request, "Requerimiento enviado correctamente.")
    else:
        messages.warning(request, "El requerimiento ya está confirmado y no puede ser enviado.")

    return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

@login_required
def detalle_requerimiento(request, requerimiento_id):
    requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    es_almacen = request.user.groups.filter(name='Almacen').exists()
    es_departamento = request.user.groups.filter(name='Departamento').exists()
    es_gestor = request.user.groups.filter(name='Gestor').exists()
    tiene_departamento_asignado = UsuarioDepartamento.objects.filter(
        usuario=request.user,
        departamento=requerimiento.departamento
    ).exists()
    solicitud_relacionada = SolicitudRequerimiento.objects.filter(requerimiento=requerimiento).select_related('usuario_solicitante').first()
    if es_gestor:
        departamentos_usuario = UsuarioDepartamento.objects.filter(usuario=request.user).values_list('departamento_id', flat=True)
        gestor_puede_ver = Requerimiento.objects.filter(
            Q(pk=requerimiento.pk),
            Q(departamento_id__in=departamentos_usuario) | Q(solicitudes_origen__usuario_solicitante=request.user)
        ).exists()
        if not gestor_puede_ver:
            messages.error(request, "No tiene permiso para ver este requerimiento.")
            return redirect('almacen:seguimiento_requerimientos')
    elif not (es_admin or es_almacen or tiene_departamento_asignado):
        messages.error(request, "No tiene permiso para realizar esta acción.")
        return redirect('almacen:acceso_denegado')

    puede_gestionar_requerimiento = (not es_gestor) and (es_admin or (es_departamento and tiene_departamento_asignado))
    puede_despachar_requerimiento = es_admin or es_almacen
    puede_anular_requerimiento = puede_despachar_requerimiento and requerimiento.estado in ['pendiente', 'enviado']

    # Asignaciones
    asignaciones_qs = (
        AsignacionDetalleFactura.objects
        .filter(destino=requerimiento.departamento)
        .values('articulo')
        .annotate(total_asignado=Sum('cantidad_asignada'))
    )
    asignaciones_dict = {item['articulo']: item['total_asignado'] for item in asignaciones_qs}

    # Despachados
    despachos_qs = (
        DetalleRequerimiento.objects
        .filter(requerimiento__departamento=requerimiento.departamento)
        .values('articulo')
        .annotate(total_despachado=Sum('cantidad_despachada'))
    )
    despachos_dict = {item['articulo']: item['total_despachado'] for item in despachos_qs}

    # Calcular stock disponible
    stock_disponible = {}
    for articulo_id, total_asignado in asignaciones_dict.items():
        total_despachado = despachos_dict.get(articulo_id, 0)
        disponible = total_asignado - total_despachado
        if disponible > 0:
            stock_disponible[articulo_id] = disponible

    # 🔁 Convertir claves del stock a str para compatibilidad con Django template
    stock_disponible = {
        str(articulo_id): disponible
        for articulo_id, disponible in stock_disponible.items()
    }

    # Filtrar artículos que tienen stock disponible
    articulos = Articulo.objects.filter(id__in=stock_disponible.keys())

    # Lista de detalles ya agregados al requerimiento
    detalles_requerimiento = DetalleRequerimiento.objects.filter(requerimiento=requerimiento)

    asignaciones_ubicacion = []
    if es_admin or es_almacen or tiene_departamento_asignado:
        for asignacion in (
            AsignacionDetalleFactura.objects
            .filter(destino=requerimiento.departamento)
            .select_related('articulo', 'destino')
            .order_by('-fecha_asignacion')
        ):
            asignaciones_ubicacion.append({
                'articulo': asignacion.articulo,
                'ubicacion': asignacion.destino,
                'cantidad_asignada': asignacion.cantidad_asignada,
                'disponible': stock_disponible.get(str(asignacion.articulo_id), 0),
                'observacion': asignacion.descripcion,
                'fecha_asignacion': asignacion.fecha_asignacion,
            })

    # Pasar motivo de rechazo al contexto si el estado es 'rechazado'
    motivo_rechazo = None
    if requerimiento.estado == 'rechazado':
        motivo_rechazo = requerimiento.motivo_rechazo

    if request.method == 'POST':
        if not puede_gestionar_requerimiento:
            messages.error(request, "No tiene permiso para modificar requerimientos.")
            return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)
        articulo_id = request.POST.get('articulo')
        cantidad = int(request.POST.get('cantidad', 0))
        observacion = request.POST.get('observaciones', '')

        if not articulo_id or cantidad <= 0:
            messages.error(request, "Debe seleccionar un artículo y una cantidad válida.")
        else:
            articulo = Articulo.objects.get(id=articulo_id)
            disponible = stock_disponible.get(str(articulo.id), 0)

            if cantidad > disponible:
                messages.error(request, f"No puedes requerir más de {disponible} unidades de {articulo.nombre}.")
            else:
                DetalleRequerimiento.objects.create(
                    requerimiento=requerimiento,
                    articulo=articulo,
                    cantidad=cantidad,
                    observacion=observacion
                )
                messages.success(request, "Detalle agregado correctamente.")
                return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

    return render(request, 'almacen/detalle_requerimiento.html', {
        'requerimiento': requerimiento,
        'articulos': articulos,
        'detalles_requerimiento': detalles_requerimiento,
        'stock_disponible': stock_disponible,
        'es_admin': es_admin,
        'es_almacen': es_almacen,
        'es_departamento': es_departamento,
        'es_gestor': es_gestor,
        'puede_gestionar_requerimiento': puede_gestionar_requerimiento,
        'puede_despachar_requerimiento': puede_despachar_requerimiento,
        'puede_anular_requerimiento': puede_anular_requerimiento,
        'puede_imprimir_requerimiento': (not es_gestor) and (es_admin or es_almacen or tiene_departamento_asignado),
        'solicitud_relacionada': solicitud_relacionada,
        'motivo_rechazo': motivo_rechazo,  # Agregar el motivo de rechazo al contexto
    })



@login_required
@require_GET
def detalle_requerimiento_api(request, detalle_id):
    if request.user.groups.filter(name='Gestor').exists() or not request.user.groups.filter(name__in=['Administrador', 'Departamento']).exists():
        return JsonResponse({'error': 'No tiene permiso para consultar esta acción administrativa.'}, status=403)
    try:
        detalle = DetalleRequerimiento.objects.get(id=detalle_id)
        data = {
            'id': detalle.id,
            'articulo_id': detalle.articulo.id,
            'cantidad': detalle.cantidad,
            'observacion': detalle.observacion,  # usa el nombre real del campo en tu modelo
        }
        return JsonResponse(data)
    except DetalleRequerimiento.DoesNotExist:
        raise Http404("Detalle no encontrado")
  
@login_required
def editar_detalle_requerimiento(request):
    if not request.user.groups.filter(name__in=['Administrador', 'Almacen', 'Departamento']).exists():
        messages.error(request, "No tiene permiso para realizar esta acción.")
        return redirect('almacen:acceso_denegado')

    if request.method == 'POST':
        detalle_id = request.POST.get('detalle_id')
        articulo_id = request.POST.get('articulo')
        cantidad = request.POST.get('cantidad')
        observaciones = request.POST.get('observaciones')

        try:
            detalle = DetalleRequerimiento.objects.get(id=detalle_id)
            articulo = Articulo.objects.get(id=articulo_id)
            detalle.articulo = articulo
            detalle.cantidad = int(cantidad)
            detalle.observacion = observaciones  # o 'observaciones' según modelo
            detalle.save()
            messages.success(request, "Detalle actualizado correctamente.")
        except (DetalleRequerimiento.DoesNotExist, Articulo.DoesNotExist):
            messages.error(request, "Error al actualizar el detalle.")
        
        return redirect('almacen:detalle_requerimiento', requerimiento_id=detalle.requerimiento.id)

    return redirect('almacen:detalle_requerimiento')  # o donde quieras    


from decimal import Decimal
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML


def exportar_kardex_pdf(request, articulo_id):
    from decimal import Decimal
    from django.shortcuts import get_object_or_404, render
    from django.http import HttpResponse
    from weasyprint import HTML

    articulo = get_object_or_404(Articulo, id=articulo_id)
    movimientos = Kardex.objects.filter(articulo=articulo).order_by('fecha', 'id')

    movimientos_con_precios = []
    total_ingresos = 0
    total_costo_ingresos = Decimal('0.00')
    total_salidas = 0
    total_costo_salidas = Decimal('0.00')

    saldo_costo_acumulado = Decimal('0.00')
    saldo_unidades = 0

    for m in movimientos:
        m.precio_unitario = Decimal('0.00')
        m.precio_total = Decimal('0.00')

        if m.tipo_movimiento == 'INGRESO' and m.fuente_factura:
            m.precio_unitario = m.fuente_factura.precio_unitario
            m.precio_total = m.fuente_factura.precio_total
            total_ingresos += m.cantidad
            total_costo_ingresos += m.precio_total

        elif m.tipo_movimiento == 'SALIDA':
            ultimo_ingreso_kardex = Kardex.objects.filter(
                articulo=articulo,
                tipo_movimiento='INGRESO',
                fecha__lt=m.fecha
            ).order_by('-fecha', '-id').first()

            if ultimo_ingreso_kardex and ultimo_ingreso_kardex.fuente_factura:
                precio_unitario = ultimo_ingreso_kardex.fuente_factura.precio_unitario
                m.precio_unitario = precio_unitario
                m.precio_total = m.cantidad * precio_unitario

            total_salidas += m.cantidad
            total_costo_salidas += m.precio_total

        if m.tipo_movimiento == 'INGRESO':
            saldo_costo_acumulado += m.precio_total
        else:
            saldo_costo_acumulado -= m.precio_total

        m.saldo_costo_total = saldo_costo_acumulado
        m.precio_unitario_saldo = m.saldo_actual and (m.saldo_costo_total / m.saldo_actual) or Decimal('0.00')

        movimientos_con_precios.append(m)

    MAX_MOVS_POR_HOJA = 7
    hojas = []
    hoja_actual = []
    linea_global = 1
    saldo_unidades = 0
    saldo_costo = Decimal('0.00')

    referencia_anterior = None  # Inicializa sin referencia anterior

    for m in movimientos_con_precios:
        if len(hoja_actual) == 0:
            hoja_actual.append({
                'tipo': 'VIENEN',
                'saldo_unidades': saldo_unidades,
                'saldo_costo_unitario': saldo_unidades and (saldo_costo / saldo_unidades) or Decimal('0.00'),
                'saldo_costo_total': saldo_costo
            })

        hoja_actual.append({
            'tipo': 'MOVIMIENTO',
            'nro': linea_global,
            'movimiento': m
        })

        saldo_unidades = m.saldo_actual
        saldo_costo = m.saldo_costo_total
        linea_global += 1

        if len(hoja_actual) == MAX_MOVS_POR_HOJA:
            numero_kardex_hoja = hoja_actual[1]['movimiento'].numero_kardex
            hojas.append({
                'numero_kardex': numero_kardex_hoja,
                'referencia_anterior': referencia_anterior,
                'lineas': hoja_actual,
            })
            referencia_anterior = numero_kardex_hoja  # Actualiza referencia para la siguiente hoja
            hoja_actual = []

    if hoja_actual:
        numero_kardex_hoja = hoja_actual[1]['movimiento'].numero_kardex
        hojas.append({
            'numero_kardex': numero_kardex_hoja,
            'referencia_anterior': referencia_anterior,
            'lineas': hoja_actual,
        })

    totales = {
        'total_ingresos': total_ingresos,
        'total_costo_ingresos': total_costo_ingresos,
        'total_salidas': total_salidas,
        'total_costo_salidas': total_costo_salidas,
        'saldo_final_unidades': total_ingresos - total_salidas,
        'saldo_final_costo': total_costo_ingresos - total_costo_salidas
    }

    html_string = render(request, 'almacen/pdf_kardex.html', {
        'articulo': articulo,
        'hojas': hojas,
        'totales': totales,
    }).content.decode('utf-8')

    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="kardex_{articulo.codigo}.pdf"'
    return response




@login_required
def exportar_requerimiento_pdf(request, requerimiento_id):
    requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    es_almacen = request.user.groups.filter(name='Almacen').exists()
    tiene_departamento_asignado = UsuarioDepartamento.objects.filter(
        usuario=request.user,
        departamento=requerimiento.departamento
    ).exists()
    if request.user.groups.filter(name='Gestor').exists() or not (es_admin or es_almacen or tiene_departamento_asignado):
        messages.error(request, "No tiene permiso para imprimir requerimientos.")
        return redirect('almacen:seguimiento_requerimientos')
    detalles = requerimiento.detalles.select_related('articulo__unidad_medida', 'articulo__categoria')
    institucion = Institucion.objects.first()  # ✅ Agregamos esto

    detalles_pdf = []
    total_general = Decimal('0.00')
    for detalle in detalles:
        cantidad = detalle.cantidad_despachada if requerimiento.estado in ['despachado', 'parcial'] and detalle.cantidad_despachada else detalle.cantidad
        cantidad_decimal = Decimal(cantidad or 0)
        ultimo_ingreso = DetalleFactura.objects.filter(articulo=detalle.articulo).order_by('-form1h__fecha_factura', '-id').first()
        # Si no hay un Formulario 1H asociado al artículo, se muestra Q0.00 sin inventar precios.
        costo_unitario = Decimal(ultimo_ingreso.precio_unitario or 0) if ultimo_ingreso else Decimal('0.00')
        total = cantidad_decimal * costo_unitario
        total_general += total
        detalles_pdf.append({
            'codigo': detalle.articulo.codigo,
            'descripcion': detalle.articulo.nombre,
            'unidad': detalle.articulo.unidad_medida or '',
            'costo_unitario': costo_unitario,
            'cantidad': cantidad,
            'total': total,
        })

    elaborado_por = requerimiento.creado_por.get_full_name() or requerimiento.creado_por.username

    html_string = render_to_string('almacen/pdf_requerimiento.html', {
        'requerimiento': requerimiento,
        'detalles': detalles_pdf,
        'total_general': total_general,
        'departamento_solicitante': requerimiento.departamento.nombre,
        'fecha': requerimiento.fecha_creacion.strftime('%d/%m/%Y'),
        'elaborado_por': elaborado_por,
        'motivo': requerimiento.motivo or 'No especificado.',
        'correlativo': f'REQ-{requerimiento.id:06d}',
        'institucion': institucion,  # ✅ Pasamos la instancia al template
    })

    # ✅ base_url ayuda a resolver rutas de imágenes y estáticos
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="requerimiento_{requerimiento.id}.pdf"'
    return response

@login_required
@grupo_requerido('Administrador', 'Almacen')
def historial_kardex_articulo(request, articulo_id):
    articulo = get_object_or_404(Articulo, id=articulo_id)
    movimientos = Kardex.objects.filter(articulo=articulo).order_by('fecha', 'id')
    return render(request, 'almacen/historial_kardex.html', {
        'articulo': articulo,
        'movimientos': movimientos
    })


def construir_stock_list_administrador():
    ingresos = DetalleFactura.objects.filter(
        form1h__estado='confirmado'
    ).values(
        'articulo__id', 'articulo__codigo', 'articulo__nombre',
        'articulo__categoria__nombre', 'articulo__renglon_presupuestario'
    ).annotate(total_ingresado=Sum('cantidad'))

    ingreso_dict = {item['articulo__id']: item for item in ingresos}
    articulo_ids = set(ingreso_dict.keys())

    asignaciones_directas = AsignacionDetalleFactura.objects.values('articulo__id').annotate(
        total_asignado=Sum('cantidad_asignada')
    )
    asignado_directo_dict = {item['articulo__id']: item['total_asignado'] or 0 for item in asignaciones_directas}
    articulo_ids.update(asignado_directo_dict.keys())

    asignaciones_departamento = AsignacionDetalleFactura.objects.values(
        'articulo__id', 'destino__nombre'
    ).annotate(total_asignado_dept=Sum('cantidad_asignada'))
    asignaciones_dept_dict = {}
    for asig in asignaciones_departamento:
        asignaciones_dept_dict.setdefault(asig['articulo__id'], []).append(
            f"{asig['destino__nombre']} ({asig['total_asignado_dept']})"
        )

    asignaciones_division = DivisionArticulo.objects.filter(activo=True).values(
        'articulo__id', 'division__nombre'
    ).annotate(total_asignado=Sum('cantidad_asignada'))
    asignado_division_dict = defaultdict(Decimal)
    divisiones_dict = {}
    for div in asignaciones_division:
        articulo_id = div['articulo__id']
        cantidad = div['total_asignado'] or Decimal('0')
        asignado_division_dict[articulo_id] += cantidad
        divisiones_dict.setdefault(articulo_id, []).append(f"{div['division__nombre']} ({cantidad})")
        articulo_ids.add(articulo_id)

    ubicaciones_division = DivisionArticuloUbicacion.objects.filter(activo=True).values(
        'division_articulo__articulo__id'
    ).annotate(
        asignado_ubicaciones=Sum('cantidad_asignada'),
        reservado=Sum('cantidad_reservada'),
        consumido=Sum('cantidad_consumida')
    )
    asignado_ubicaciones_dict = defaultdict(Decimal)
    reservado_dict = defaultdict(Decimal)
    consumido_dict = defaultdict(Decimal)
    for item in ubicaciones_division:
        articulo_id = item['division_articulo__articulo__id']
        asignado_ubicaciones_dict[articulo_id] = item['asignado_ubicaciones'] or Decimal('0')
        reservado_dict[articulo_id] = item['reservado'] or Decimal('0')
        consumido_dict[articulo_id] = item['consumido'] or Decimal('0')
        articulo_ids.add(articulo_id)

    faltantes = articulo_ids - set(ingreso_dict.keys())
    if faltantes:
        for articulo in Articulo.objects.filter(id__in=faltantes).select_related('categoria'):
            ingreso_dict[articulo.id] = {
                'articulo__id': articulo.id,
                'articulo__codigo': articulo.codigo,
                'articulo__nombre': articulo.nombre,
                'articulo__categoria__nombre': articulo.categoria.nombre if articulo.categoria else None,
                'articulo__renglon_presupuestario': articulo.renglon_presupuestario,
                'total_ingresado': 0,
            }

    stock_list = []
    for articulo_id, item in ingreso_dict.items():
        total_ingresado = item.get('total_ingresado') or 0
        asignado_directo = asignado_directo_dict.get(articulo_id, 0)
        asignado_divisiones = asignado_division_dict.get(articulo_id, Decimal('0'))
        asignado_ubicaciones = asignado_ubicaciones_dict.get(articulo_id, Decimal('0'))
        reservado_total = reservado_dict.get(articulo_id, Decimal('0'))
        consumido_total = consumido_dict.get(articulo_id, Decimal('0'))
        asignado_total = asignado_directo + asignado_divisiones
        disponible_real = total_ingresado - asignado_total
        requiere_revision = disponible_real < 0
        disponible_general = max(0, disponible_real)
        departamentos = ", ".join(asignaciones_dept_dict.get(articulo_id, [])) or "Sin asignar"
        divisiones = ", ".join(divisiones_dict.get(articulo_id, [])) or "Sin división"

        stock_list.append({
            'articulo_id': articulo_id,
            'articulo_codigo': item.get('articulo__codigo'),
            'renglon_presupuestario': item.get('articulo__renglon_presupuestario'),
            'categoria': item.get('articulo__categoria__nombre'),
            'articulo': item['articulo__nombre'],
            'ingresado': total_ingresado,
            'asignado': asignado_total,
            'asignado_directo': asignado_directo,
            'asignado_divisiones': asignado_divisiones,
            'asignado_ubicaciones_division': asignado_ubicaciones,
            'reservado_total': reservado_total,
            'consumido_total': consumido_total,
            'disponible': disponible_general,
            'requiere_revision': requiere_revision,
            'departamentos': departamentos,
            'divisiones': divisiones,
        })
    return stock_list


@login_required
@grupo_requerido('Administrador', 'Almacen')
def ver_stock_formulario_1h(request):
    context = {
        'stock_list': construir_stock_list_administrador(),
        'departamentos': Departamento.objects.all(),
        'es_administrador': request.user.is_superuser or request.user.groups.filter(name='Administrador').exists(),
    }
    return render(request, 'almacen/stock_formulario_1h.html', context)
    

@login_required
@grupo_requerido('Administrador', 'Almacen')
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def asignar_departamento_usuario(request):
    if request.method == 'POST':
        form = UsuarioDepartamentoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Departamento asignado correctamente al usuario.')
                return redirect('almacen:asignar_departamento')
            except:
                messages.error(request, 'Este usuario ya está asignado a ese departamento.')
    else:
        form = UsuarioDepartamentoForm()

    # Agrupar departamentos por usuario
    asignaciones = UsuarioDepartamento.objects.select_related('usuario', 'departamento')
    usuarios_con_departamentos = defaultdict(list)
    for asignacion in asignaciones:
        usuarios_con_departamentos[asignacion.usuario].append(asignacion.departamento)

    context = {
        'form': form,
        'usuarios_con_departamentos': usuarios_con_departamentos.items(),
    }
    return render(request, 'almacen/asignar_departamento.html', context)

def eliminar_asignacion(request, usuario_id, departamento_id):
    if request.method == 'POST':
        asignacion = get_object_or_404(UsuarioDepartamento, usuario_id=usuario_id, departamento_id=departamento_id)
        asignacion.delete()
        messages.success(request, 'Asignación eliminada correctamente.')
    else:
        messages.error(request, 'Método no permitido.')
    return redirect('almacen:asignar_departamento')


@login_required
def lista_departamentos(request):
    es_departamento = request.user.groups.filter(name='Departamento').exists()
    es_gestor = request.user.groups.filter(name='Gestor').exists()

    if es_departamento or es_gestor:
        # Obtener todos los objetos UsuarioDepartamento vinculados al usuario
        usuario_departamentos = UsuarioDepartamento.objects.filter(usuario=request.user)
        # Mostrar todos los departamentos asociados a esas instancias
        departamentos = Departamento.objects.filter(usuariodepartamento__in=usuario_departamentos)
    else:
        departamentos = Departamento.objects.all()

    return render(request, 'almacen/lista_departamentos.html', {
        'departamentos': departamentos,
        'es_departamento': es_departamento
    })

def acceso_denegado(request, exception=None):
    return render(request, 'almacen/403.html', status=403)

from django.db.models import Q


def construir_resumen_gestor_departamento(departamento):
    detalle_asignaciones = []

    despachados = (
        DetalleRequerimiento.objects
        .filter(requerimiento__departamento=departamento)
        .values('articulo__id')
        .annotate(total_despachado=Sum('cantidad_despachada'))
    )
    despachados_dict = {d['articulo__id']: Decimal(d['total_despachado'] or 0) for d in despachados}

    division_consumido_por_articulo = defaultdict(Decimal)
    asignaciones_division = (
        DivisionArticuloUbicacion.objects
        .filter(
            ubicacion=departamento,
            activo=True,
            division_articulo__activo=True,
            division_articulo__division__activa=True,
        )
        .select_related('division_articulo__articulo', 'division_articulo__articulo__categoria', 'division_articulo__division')
        .order_by('division_articulo__division__nombre', 'division_articulo__articulo__nombre')
    )
    for asignacion in asignaciones_division:
        articulo = asignacion.division_articulo.articulo
        cantidad_asignada = Decimal(asignacion.cantidad_asignada or 0)
        cantidad_reservada = Decimal(asignacion.cantidad_reservada or 0)
        cantidad_consumida = Decimal(asignacion.cantidad_consumida or 0)
        disponible_real = cantidad_asignada - cantidad_reservada - cantidad_consumida
        division_consumido_por_articulo[articulo.id] += cantidad_consumida
        detalle_asignaciones.append({
            'articulo': articulo,
            'origen': 'División',
            'division': asignacion.division_articulo.division.nombre,
            'departamento': departamento,
            'cantidad_asignada': cantidad_asignada,
            'reservado': cantidad_reservada,
            'consumido': cantidad_consumida,
            'disponible': max(Decimal('0'), disponible_real),
            'requiere_revision': disponible_real < 0,
            'fecha': asignacion.fecha_asignacion,
            'descripcion': asignacion.division_articulo.observacion,
            'transferible': False,
        })

    consumo_directo_restante = defaultdict(Decimal)
    for articulo_id, total_despachado in despachados_dict.items():
        consumo_directo_restante[articulo_id] = max(
            Decimal('0'),
            total_despachado - division_consumido_por_articulo.get(articulo_id, Decimal('0'))
        )

    directas = (
        AsignacionDetalleFactura.objects
        .filter(destino=departamento)
        .select_related('articulo', 'articulo__categoria')
        .order_by('-fecha_asignacion')
    )
    for asignacion in directas:
        articulo = asignacion.articulo
        cantidad_asignada = Decimal(asignacion.cantidad_asignada or 0)
        cantidad_reservada = Decimal('0')
        cantidad_consumida = min(cantidad_asignada, consumo_directo_restante.get(articulo.id, Decimal('0')))
        consumo_directo_restante[articulo.id] -= cantidad_consumida
        disponible_real = cantidad_asignada - cantidad_reservada - cantidad_consumida
        detalle_asignaciones.append({
            'articulo': articulo,
            'origen': 'Directo',
            'division': None,
            'departamento': departamento,
            'cantidad_asignada': cantidad_asignada,
            'reservado': cantidad_reservada,
            'consumido': cantidad_consumida,
            'disponible': max(Decimal('0'), disponible_real),
            'requiere_revision': disponible_real < 0,
            'fecha': asignacion.fecha_asignacion,
            'descripcion': asignacion.descripcion,
            'transferible': True,
        })

    resumen_por_articulo = {}
    for fila in detalle_asignaciones:
        articulo = fila['articulo']
        item = resumen_por_articulo.setdefault(articulo.id, {
            'articulo_id': articulo.id,
            'articulo_codigo': articulo.codigo,
            'renglon_presupuestario': articulo.renglon_presupuestario,
            'nombre_articulo': articulo.nombre,
            'categoria': articulo.categoria.nombre if articulo.categoria else None,
            'asignado_directo': Decimal('0'),
            'asignado_division': Decimal('0'),
            'reservado': Decimal('0'),
            'despachado': Decimal('0'),
            'disponible': Decimal('0'),
            'requiere_revision': False,
        })
        if fila['origen'] == 'Directo':
            item['asignado_directo'] += fila['cantidad_asignada']
        else:
            item['asignado_division'] += fila['cantidad_asignada']
        item['reservado'] += fila['reservado']
        item['despachado'] += fila['consumido']
        item['disponible'] += fila['disponible']
        item['requiere_revision'] = item['requiere_revision'] or fila['requiere_revision']

    resumen_stock = []
    for item in resumen_por_articulo.values():
        item['asignado'] = item['asignado_directo'] + item['asignado_division']
        resumen_stock.append(item)

    resumen_stock.sort(key=lambda item: item['nombre_articulo'])
    detalle_asignaciones.sort(key=lambda item: (item['articulo'].nombre, item['origen'], item['division'] or ''))
    return resumen_stock, detalle_asignaciones

@login_required
def detalle_departamento(request, pk):
    departamento = get_object_or_404(Departamento, pk=pk)

    es_departamento = request.user.groups.filter(name='Departamento').exists()
    es_gestor = request.user.groups.filter(name='Gestor').exists()
    es_admin = request.user.groups.filter(name='Administrador').exists()

    tiene_acceso = es_admin or UsuarioDepartamento.objects.filter(usuario=request.user, departamento=departamento).exists()

    asignaciones_agrupadas = []
    asignaciones_detalle = []
    resumen_stock = []
    departamentos = []
    historial_transferencias = []

    if tiene_acceso:
        departamentos = Departamento.objects.exclude(id=departamento.id)

        resumen_stock, asignaciones_detalle = construir_resumen_gestor_departamento(departamento)
        asignaciones_agrupadas = resumen_stock

        # Historial de transferencias donde el departamento es origen o destino
        historial_transferencias = (
            HistorialTransferencia.objects
            .filter(Q(departamento_origen=departamento) | Q(departamento_destino=departamento))
            .order_by('-fecha_transferencia')
        )

    return render(request, 'almacen/detalle_departamento.html', {
        'departamento': departamento,
        'asignaciones_agrupadas': asignaciones_agrupadas,
        'asignaciones_detalle': asignaciones_detalle,
        'resumen_stock': resumen_stock,
        'tiene_acceso': tiene_acceso,
        'es_departamento': es_departamento,
        'es_gestor': es_gestor,
        'es_administrador': es_admin,
        'departamentos': departamentos,
        'historial_transferencias': historial_transferencias,
    })



def calcular_disponibilidad_para_division(articulo):
    total_ingresado = DetalleFactura.objects.filter(
        articulo=articulo,
        form1h__estado='confirmado'
    ).aggregate(total=Sum('cantidad'))['total'] or 0
    asignado_directo = AsignacionDetalleFactura.objects.filter(
        articulo=articulo
    ).aggregate(total=Sum('cantidad_asignada'))['total'] or 0
    asignado_divisiones = DivisionArticulo.objects.filter(
        articulo=articulo,
        activo=True
    ).aggregate(total=Sum('cantidad_asignada'))['total'] or Decimal('0')
    disponible_real = Decimal(total_ingresado) - Decimal(asignado_directo) - Decimal(asignado_divisiones)
    return {
        'total_ingresado': total_ingresado,
        'asignado_directo': asignado_directo,
        'asignado_divisiones': asignado_divisiones,
        'disponible_para_division': max(Decimal('0'), disponible_real),
        'requiere_revision': disponible_real < 0,
    }


@login_required
@grupo_requerido('Administrador', 'Almacen')
def asignar_articulo_division(request, articulo_id):
    articulo = get_object_or_404(Articulo.objects.select_related('categoria'), pk=articulo_id)
    disponibilidad = calcular_disponibilidad_para_division(articulo)
    divisiones = DivisionAlmacen.objects.filter(activa=True).order_by('nombre')

    if request.method == 'POST':
        division_id = request.POST.get('division')
        cantidad_raw = request.POST.get('cantidad')
        division = divisiones.filter(pk=division_id).first()

        try:
            cantidad = Decimal(cantidad_raw or '0')
        except Exception:
            cantidad = Decimal('0')

        if not division:
            messages.error(request, 'Debe seleccionar una división activa.')
        elif cantidad <= 0:
            messages.error(request, 'La cantidad a asignar debe ser mayor a cero.')
        elif disponibilidad['disponible_para_division'] <= 0:
            messages.error(request, 'No hay disponibilidad para asignar este artículo a divisiones.')
        elif cantidad > disponibilidad['disponible_para_division']:
            messages.error(request, 'No puede asignar más de la cantidad disponible para divisiones.')
        else:
            try:
                with transaction.atomic():
                    division_articulo = DivisionArticulo.objects.select_for_update().filter(
                        division=division,
                        articulo=articulo,
                        activo=True
                    ).first()
                    if division_articulo:
                        division_articulo.cantidad_asignada += cantidad
                        division_articulo.asignado_por = request.user
                    else:
                        division_articulo = DivisionArticulo(
                            division=division,
                            articulo=articulo,
                            cantidad_asignada=cantidad,
                            asignado_por=request.user,
                            activo=True,
                        )
                    division_articulo.save()
                messages.success(request, 'Artículo asignado a la división correctamente.')
                return redirect('almacen:ver_stock_formulario_1h')
            except ValidationError as e:
                messages.error(request, e.message_dict if hasattr(e, 'message_dict') else e.messages)

        disponibilidad = calcular_disponibilidad_para_division(articulo)

    return render(request, 'almacen/asignar_articulo_division.html', {
        'articulo': articulo,
        'divisiones': divisiones,
        'disponibilidad': disponibilidad,
    })

@login_required
@grupo_requerido('Administrador', 'Almacen')
def transferir_articulos(request):
    departamentos = Departamento.objects.all()

    if request.method == 'POST':
        departamento_origen_id = request.POST.get('departamento_origen')
        articulo_id = request.POST.get('articulo')
        cantidad_str = request.POST.get('cantidad_transferir')
        departamento_destino_id = request.POST.get('departamento_destino')
        observacion = request.POST.get('observacion', '').strip()  # <-- aquí

        if not (departamento_origen_id and articulo_id and cantidad_str and departamento_destino_id):
            messages.error(request, 'Faltan datos para realizar la transferencia.')
            return redirect('almacen:lista_departamentos')

        try:
            cantidad = int(cantidad_str)
        except ValueError:
            messages.error(request, 'Cantidad inválida.')
            return redirect('almacen:detalle_departamento', pk=departamento_origen_id)

        departamento_origen = get_object_or_404(Departamento, id=departamento_origen_id)
        articulo = get_object_or_404(Articulo, id=articulo_id)
        departamento_destino = get_object_or_404(Departamento, id=departamento_destino_id)

        asignacion_origen = AsignacionDetalleFactura.objects.filter(
            articulo=articulo, destino=departamento_origen
        ).first()

        if not asignacion_origen or asignacion_origen.cantidad_asignada < cantidad:
            messages.error(request, 'No hay suficiente cantidad en el departamento de origen.')
            return redirect('almacen:detalle_departamento', pk=departamento_origen_id)

        # Reducir cantidad en origen
        asignacion_origen.cantidad_asignada -= cantidad
        asignacion_origen.save()

        # Incrementar cantidad en destino, o crear asignación si no existe
        asignacion_destino, created = AsignacionDetalleFactura.objects.get_or_create(
            articulo=articulo, destino=departamento_destino,
            defaults={'cantidad_asignada': 0}
        )
        asignacion_destino.cantidad_asignada += cantidad
        asignacion_destino.save()

        # Guardar historial de transferencia con observación
        HistorialTransferencia.objects.create(
            articulo=articulo,
            cantidad=cantidad,
            departamento_origen=departamento_origen,
            departamento_destino=departamento_destino,
            usuario=request.user,
            observacion=observacion
        )

        messages.success(request, f'Se transfirieron {cantidad} unidades de {articulo.nombre} de {departamento_origen.nombre} a {departamento_destino.nombre}.')

        return redirect('almacen:detalle_departamento', pk=departamento_origen_id)

    return redirect('almacen:lista_departamentos')


@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_asignacion_detalle(request):
    if request.method == 'POST':
        form = AsignacionDetalleFacturaForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Asignación creada correctamente.")
                return redirect('almacen:crear_asignacion_detalle')
            except Exception as e:
                messages.error(request, f"Error al asignar: {e}")
    else:
        form = AsignacionDetalleFacturaForm()

    # 👇 Estas líneas deben estar bien indentadas dentro de la función
    stock_por_articulo = DetalleFactura.objects.filter(
        form1h__estado='confirmado'
    ).values('articulo').annotate(stock_total=Sum('cantidad'))

    asignado_por_articulo = AsignacionDetalleFactura.objects.values('articulo').annotate(
        asignado_total=Sum('cantidad_asignada'))

    stock_dict = {}
    for item in stock_por_articulo:
        articulo_id = item['articulo']
        total_stock = item['stock_total'] or 0
        total_asignado = next((a['asignado_total'] for a in asignado_por_articulo if a['articulo'] == articulo_id), 0)
        stock_disponible = total_stock - (total_asignado or 0)
        stock_dict[articulo_id] = stock_disponible

    ultima_asignacion = AsignacionDetalleFactura.objects.order_by('-fecha_asignacion').first()

    return render(request, 'almacen/crear_asignacion_detalle.html', {
        'form': form,
        'stock_dict': stock_dict,
        'ultima_asignacion': ultima_asignacion,
    })

   
@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_asignacion_detalle_articulo(request):
    if request.method == 'POST':
        form = AsignacionDetalleFacturaForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Asignación creada correctamente.")
                return redirect('almacen:crear_asignacion_detalle_articulo')
            except Exception as e:
                messages.error(request, f"Error al asignar: {e}")
    else:
        form = AsignacionDetalleFacturaForm()

    # Calcular stock ingresado
    ingresos = DetalleFactura.objects.filter(
        form1h__estado='confirmado'
    ).values('articulo__id', 'articulo__codigo', 'articulo__nombre', 'articulo__categoria__nombre', 'articulo__renglon_presupuestario').annotate(
        total_ingresado=Sum('cantidad')
    )

    ingreso_dict = {item['articulo__id']: item for item in ingresos}

    # Asignaciones totales
    asignaciones = AsignacionDetalleFactura.objects.values('articulo__id').annotate(
        total_asignado=Sum('cantidad_asignada')
    )

    # Asignaciones por departamento
    asignaciones_departamento = AsignacionDetalleFactura.objects.values(
        'articulo__id', 'destino__nombre'
    ).annotate(total_asignado_dept=Sum('cantidad_asignada'))

    asignaciones_dept_dict = {}
    for asig in asignaciones_departamento:
        art_id = asig['articulo__id']
        depto_nombre = asig['destino__nombre']
        cantidad = asig['total_asignado_dept']
        if art_id not in asignaciones_dept_dict:
            asignaciones_dept_dict[art_id] = []
        asignaciones_dept_dict[art_id].append(f"{depto_nombre} ({cantidad})")

    # Combinar ingresos y asignaciones
    for item in asignaciones:
        articulo_id = item['articulo__id']
        if articulo_id in ingreso_dict:
            ingreso_dict[articulo_id]['total_asignado'] = item['total_asignado']
        else:
            articulo = Articulo.objects.get(id=articulo_id)
            ingreso_dict[articulo_id] = {
                'articulo__id': articulo_id,
                'articulo__codigo': articulo.codigo,
                'articulo__nombre': articulo.nombre,
                'articulo__categoria__nombre': articulo.categoria.nombre if articulo.categoria else None,
                'articulo__renglon_presupuestario': articulo.renglon_presupuestario,
                'total_ingresado': 0,
                'total_asignado': item['total_asignado'],
            }

    # Construir stock_list
    stock_list = []
    for item in ingreso_dict.values():
        total_ingresado = item.get('total_ingresado') or 0
        total_asignado = item.get('total_asignado') or 0
        disponible = total_ingresado - total_asignado
        departamentos = ", ".join(asignaciones_dept_dict.get(item['articulo__id'], [])) or "Sin asignar"

        stock_list.append({
            'articulo_id': item['articulo__id'],
            'articulo_codigo': item.get('articulo__codigo'),
            'renglon_presupuestario': item.get('articulo__renglon_presupuestario'),
            'categoria': item.get('articulo__categoria__nombre'),
            'articulo': item['articulo__nombre'],
            'ingresado': total_ingresado,
            'asignado': total_asignado,
            'disponible': disponible,
            'departamentos': departamentos,
        })

    context = {
        'form': form,
        'stock_list': construir_stock_list_administrador(),
        'departamentos': Departamento.objects.all(),
        'es_administrador': request.user.is_superuser or request.user.groups.filter(name='Administrador').exists(),
    }

    return render(request, 'almacen/stock_formulario_1h.html', context)


    
def buscar_articulos(request):
    term = request.GET.get('q', '')
    articulos = Articulo.objects.filter(Q(codigo__icontains=term) | Q(nombre__icontains=term))[:10]
    results = [
        {
            'id': art.id,
            'codigo': art.codigo,
            'nombre': art.nombre,
            'text': f"{art.codigo} - {art.nombre} — Categoría: {art.categoria.nombre if art.categoria else 'S/C'} — Renglón: {art.renglon_presupuestario or 'S/R'}",
            'categoria': art.categoria.nombre if art.categoria else 'S/C',
            'renglon_presupuestario': art.renglon_presupuestario,
        }
        for art in articulos
    ]
    return JsonResponse(results, safe=False)
    
@login_required
@grupo_requerido('Administrador', 'Almacen')   
def serie_form_list(request, pk=None):
    if pk:
        instance = get_object_or_404(Serie, pk=pk)
    else:
        instance = None

    if request.method == 'POST':
        form = SerieForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('almacen:lista_series')
    else:
        form = SerieForm(instance=instance)

    series = Serie.objects.all()
    return render(request, 'almacen/serie_list.html', {
        'form': form,
        'series': series,
    })

@login_required
@grupo_requerido('Administrador', 'Almacen')
@require_POST
def confirmar_form1h(request, form1h_id):
    formulario = get_object_or_404(form1h, id=form1h_id)

    if formulario.estado == 'borrador':
        formulario.estado = 'confirmado'
        formulario.save()
        messages.success(request, f'El formulario Serie {formulario.serie.serie} {formulario.numero_serie}  ha sido confirmado exitosamente.')
    elif formulario.estado == 'confirmado':
        messages.info(request, f'El formulario Serie {formulario.serie.serie} {formulario.numero_serie}  ya está confirmado.')
    else:
        messages.warning(request, f'El formulario Serie {formulario.serie.serie} {formulario.numero_serie}  no se puede confirmar en su estado actual.')

    return redirect('almacen:agregar_detalle_factura', form1h_id=form1h_id)

@login_required
@grupo_requerido('Administrador', 'Almacen')
@transaction.atomic
def anular_requerimiento(request, requerimiento_id):
    if not request.user.groups.filter(name__in=['Administrador', 'Almacen']).exists():
        messages.error(request, "No tienes permiso para anular requerimientos.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento_id)

    requerimiento = get_object_or_404(Requerimiento, id=requerimiento_id)

    if requerimiento.estado not in ['pendiente', 'enviado']:
        messages.warning(request, "Solo se pueden anular requerimientos en estado 'pendiente' o 'enviado'.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)


    if request.method == 'POST':
        motivo_anulacion = request.POST.get('motivo_anulacion')

        # Cambiar el estado del requerimiento a "rechazado" y registrar el motivo de anulacion
        requerimiento.estado = 'rechazado'
        requerimiento.motivo_rechazo = motivo_anulacion  # Usar el nuevo campo 'motivo_rechazo'
        requerimiento.fecha_rechazado = timezone.now()
        requerimiento.rechazado_por = request.user
        requerimiento.save()
        solicitud_origen = requerimiento.solicitudes_origen.first()
        if solicitud_origen:
            for detalle in solicitud_origen.detalles.select_related('division_articulo_ubicacion'):
                if detalle.division_articulo_ubicacion_id:
                    asignacion = detalle.division_articulo_ubicacion
                    asignacion.cantidad_reservada = max(Decimal('0'), asignacion.cantidad_reservada - Decimal(detalle.cantidad))
                    asignacion.save(update_fields=['cantidad_reservada', 'fecha_actualizacion'])

        messages.success(request, "Requerimiento anulado exitosamente.")
        return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)

    messages.error(request, "Método no permitido.")
    return redirect('almacen:detalle_requerimiento', requerimiento_id=requerimiento.id)



@login_required
@grupo_requerido('Administrador', 'Almacen')
@require_POST
def anular_form1h(request, form1h_id):
    formulario = get_object_or_404(form1h, id=form1h_id)

    if formulario.estado != 'anulado':
        formulario.estado = 'anulado'
        formulario.save()
        messages.success(request, f'El formulario Serie {formulario.serie.serie} {formulario.numero_serie} ha sido anulado exitosamente.')
    else:
        messages.info(request, f'El formulario Serie {formulario.serie.serie} {formulario.numero_serie} ya está anulado.')

    # Redirige a la misma vista de agregar detalles, como con "confirmar"
    return redirect('almacen:agregar_detalle_factura', form1h_id=form1h_id)


def buscar_proveedor_nit(request, nit):
    try:
        proveedor = Proveedor.objects.get(nit=nit)
        data = {
            'nombre': proveedor.nombre,
            'telefono': proveedor.telefono,
            'direccion': proveedor.direccion,
            'proveedor_id': proveedor.id,
        }
        return JsonResponse(data)
    except Proveedor.DoesNotExist:
        return JsonResponse({'error': 'Proveedor no encontrado'}, status=404)
    
def buscar_proveedor_id(request, proveedor_id):
    try:
        proveedor = Proveedor.objects.get(id=proveedor_id)
        data = {
            'nombre': proveedor.nombre,
            'telefono': proveedor.telefono,
            'direccion': proveedor.direccion,
            'nit': proveedor.nit,
        }
        return JsonResponse(data)
    except Proveedor.DoesNotExist:
        return JsonResponse({'error': 'Proveedor no encontrado'}, status=404)

@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_form1h(request):
    form1h_list = form1h.objects.all()
    form = Form1hForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            try:
                cantidad = form.cleaned_data.get('cantidad_detalles')
                nuevo_formulario = form.save()

                # Obtener o crear el contador global
                contador, _ = ContadorDetalleFactura.objects.get_or_create(id=1)

                for _ in range(cantidad):
                    if LineaLibre.objects.exists():
                        libre = LineaLibre.objects.first()
                        numero_linea = libre.id_linea
                        libre.delete()
                    else:
                        numero_linea = contador.contador
                        contador.contador += 1

                    # Asegúrate de que no esté duplicado
                    if not LineaReservada.objects.filter(numero_linea=numero_linea).exists():
                        LineaReservada.objects.create(
                            form1h=nuevo_formulario,
                            numero_linea=numero_linea,
                            disponible=True
                        )

                contador.save()

                messages.success(request, f"Formulario creado y se reservaron {cantidad} líneas.")
                return redirect('almacen:agregar_detalle_factura', form1h_id=nuevo_formulario.id)
            except ValidationError as e:
                messages.error(request, e.message)

    return render(request, 'almacen/crear_form1h.html', {
        'form': form,
        'form1h_list': form1h_list
    })

@login_required
@grupo_requerido('Administrador', 'Almacen')
def agregar_detalle_factura(request, form1h_id):
    form1h_instance = get_object_or_404(form1h, id=form1h_id)
    detalles_factura = DetalleFactura.objects.filter(form1h=form1h_instance)

    total_factura = form1h_instance.calcular_total_factura()
    articulos = Articulo.objects.all()
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.all()
    unidades = UnidadDeMedida.objects.all()

    # Filtrar líneas disponibles para este form1h
    lineas_reservadas = LineaReservada.objects.filter(
        form1h=form1h_instance,
        disponible=True
    ).order_by('numero_linea')

    print("======= LÍNEAS RESERVADAS =======")
    for linea in lineas_reservadas:
        print(f"Línea: {linea.numero_linea} | Disponible: {linea.disponible} | Formulario ID: {linea.form1h_id}")

    if request.method == "POST":
        numero_linea = request.POST.get('detalle_numero_linea')
        renglon = request.POST.get('renglon')

        # Obtener listas de folio y nomenclatura
        folios = request.POST.getlist('folio_inventario[]')
        nomenclaturas = request.POST.getlist('nomenclatura[]')

        # Clonar POST y añadir el id_linea que necesita el form
        post_data = request.POST.copy()
        post_data['id_linea'] = numero_linea  # Esto se inyecta en el form

        form = DetalleFacturaForm(post_data, form1h_instance=form1h_instance)

        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.form1h = form1h_instance
            detalle.renglon = renglon  # Puedes usar directamente: detalle.renglon = numero_linea

            try:
                linea_reservada = LineaReservada.objects.get(
                    numero_linea=numero_linea,
                    form1h=form1h_instance,
                    disponible=True
                )

                detalle.save()

                # Guardar InventarioDetalle solo si folios o nomenclaturas existen
                for folio, nomen in zip(folios, nomenclaturas):
                    if folio.strip() or nomen.strip():
                        InventarioDetalle.objects.create(
                            detalle_factura=detalle,
                            folio_inventario=folio.strip(),
                            nomenclatura=nomen.strip()
                        )

                # Marcar la línea como no disponible
                linea_reservada.disponible = False
                linea_reservada.save()

                messages.success(request, f"Detalle agregado usando línea #{numero_linea}.")
                return redirect('almacen:agregar_detalle_factura', form1h_id=form1h_id)

            except LineaReservada.DoesNotExist:
                messages.error(request, "La línea seleccionada no está disponible o ya fue utilizada.")
        else:
            print("Errores del formulario:", form.errors)
            messages.error(request, "Error al guardar el detalle. Verifica los campos.")
    else:
        form = DetalleFacturaForm(form1h_instance=form1h_instance)

    return render(request, 'almacen/agregar_detalle_factura.html', {
        'form1h_instance': form1h_instance,
        'form': form,
        'detalles_factura': detalles_factura,
        'total_factura': total_factura,
        'articulos': articulos,
        'categorias': categorias,
        'ubicaciones': ubicaciones,
        'unidades': unidades,
        'lineas_reservadas': lineas_reservadas,
    })




@login_required
@require_POST
@grupo_requerido('Administrador', 'Almacen', 'Departamento')
def eliminar_detalle_requerimiento(request, pk):
    if request.user.groups.filter(name='Gestor').exists():
        return JsonResponse({'error': 'No tiene permiso para eliminar detalles.'}, status=403)
    if request.method == 'POST':
        try:
            detalle = DetalleRequerimiento.objects.get(pk=pk)
            detalle.delete()
            return JsonResponse({'success': True})
        except DetalleRequerimiento.DoesNotExist:
            return JsonResponse({'error': 'Detalle no encontrado'}, status=404)
    return HttpResponseNotAllowed(['POST'])

# Views for Departamento
@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_departamento(request):
    departamentos = Departamento.objects.all()  # Obtener todos los departamentos
    form = DepartamentoForm(request.POST or None)  # Crear el formulario
    if form.is_valid():
        form.save()  # Guardar el nuevo departamento
        return redirect('almacen:crear_departamento')  # Redirige a la misma página para mostrar el nuevo departamento
    return render(request, 'almacen/crear_departamento.html', {'form': form, 'departamentos': departamentos})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_departamento(request, pk):
    departamento = get_object_or_404(Departamento, pk=pk)  # Obtener el departamento por su PK
    form = DepartamentoForm(request.POST or None, instance=departamento)  # Rellenar el formulario con los datos existentes
    if form.is_valid():
        form.save()  # Guardar los cambios en el departamento
        return redirect('almacen:crear_departamento')  # Redirige a la vista de creación (o a donde desees)
    return render(request, 'almacen/editar_departamento.html', {'form': form, 'departamentos': Departamento.objects.all()})


# Create your views here.
@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_articulo(request):
    articulos = Articulo.objects.all()  # Obtener todos los artículos
    form = ArticuloForm(request.POST or None)  # Crear el formulario
    if form.is_valid():
        form.save()  # Guardar el nuevo artículo
        return redirect('almacen:crear_articulo')  # Redirige a la misma página para mostrar el nuevo artículo
    return render(request, 'almacen/crear_articulo.html', {'form': form, 'articulos': articulos})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_articulo(request, pk):
    articulo = get_object_or_404(Articulo, pk=pk)  # Obtener el artículo por su PK
    form = ArticuloForm(request.POST or None, instance=articulo)  # Rellenar el formulario con los datos existentes
    if form.is_valid():
        form.save()  # Guardar los cambios en el artículo
        return redirect('almacen:crear_articulo')  # Redirige a la vista de creación (o a donde desees)
    return render(request, 'almacen/editar_articulo.html', {'form': form, 'articulos': Articulo.objects.all()})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_proveedor(request):
    proveedores = Proveedor.objects.all()  # Obtener todos los proveedores
    form = ProveedorForm(request.POST or None)  # Crear el formulario
    if form.is_valid():
        form.save()  # Guardar el nuevo proveedor
        return redirect('almacen:crear_proveedor')  # Redirige a la misma página para mostrar el nuevo proveedor
    return render(request, 'almacen/crear_proveedor.html', {'form': form, 'proveedores': proveedores})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)  # Obtener el proveedor por su PK
    form = ProveedorForm(request.POST or None, instance=proveedor)  # Rellenar el formulario con los datos existentes
    if form.is_valid():
        form.save()  # Guardar los cambios en el proveedor
        return redirect('almacen:crear_proveedor')  # Redirige a la vista de creación (o a donde desees)
    return render(request, 'almacen/editar_proveedor.html', {'form': form, 'proveedores': Proveedor.objects.all()})

# Views for Categoria
@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_categoria(request):
    categorias = Categoria.objects.all()  # Obtener todas las categorías
    form = CategoriaForm(request.POST or None)  # Crear el formulario
    if form.is_valid():
        form.save()  # Guardar la nueva categoría
        return redirect('almacen:crear_categoria')  # Redirige a la misma página para mostrar la nueva categoría
    return render(request, 'almacen/crear_categoria.html', {'form': form, 'categorias': categorias})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)  # Obtener la categoría por su PK
    form = CategoriaForm(request.POST or None, instance=categoria)  # Rellenar el formulario con los datos existentes
    if form.is_valid():
        form.save()  # Guardar los cambios en la categoría
        return redirect('almacen:crear_categoria')  # Redirige a la vista de creación (o a donde desees)
    return render(request, 'almacen/editar_categoria.html', {'form': form, 'categorias': Categoria.objects.all()})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_unidad(request):
    unidades = UnidadDeMedida.objects.all()  # Obtener todas las unidades de medida
    form = UnidadDeMedidaForm(request.POST or None)  # Crear el formulario
    if form.is_valid():
        form.save()  # Guardar la nueva unidad de medida
        return redirect('almacen:crear_unidad')  # Redirige a la misma página para mostrar la nueva unidad
    return render(request, 'almacen/crear_unidad.html', {'form': form, 'unidades': unidades})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_unidad(request, pk):
    unidad = get_object_or_404(UnidadDeMedida, pk=pk)
    if request.method == 'POST':
        form = UnidadDeMedidaForm(request.POST, instance=unidad)
        if form.is_valid():
            form.save()
            return redirect('almacen:crear_unidad')  # Redirige a la lista de unidades
    else:
        form = UnidadDeMedidaForm(instance=unidad)

    return render(request, 'almacen/crear_unidad.html', {'form': form})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_programa(request):
    programas = Programa.objects.all()
    form = ProgramaForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('almacen:crear_programa')  # Ajusta a tu namespace de URLs
    return render(request, 'almacen/crear_programa.html', {'form': form, 'programas': programas})


@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_programa(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    programas = Programa.objects.all()
    if request.method == 'POST':
        form = ProgramaForm(request.POST, instance=programa)
        if form.is_valid():
            form.save()
            return redirect('almacen:crear_programa')
    else:
        form = ProgramaForm(instance=programa)
    return render(request, 'almacen/crear_programa.html', {'form': form, 'programas': programas})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_dependencia(request):
    dependencias = Dependencia.objects.all()
    form = DependenciaForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('almacen:crear_dependencia')
    return render(request, 'almacen/crear_dependencia.html', {'form': form, 'dependencias': dependencias})


@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_dependencia(request, pk):
    dependencia = get_object_or_404(Dependencia, pk=pk)
    dependencias = Dependencia.objects.all()

    if request.method == 'POST':
        form = DependenciaForm(request.POST, instance=dependencia)
        if form.is_valid():
            form.save()
            return redirect('almacen:crear_dependencia')
    else:
        form = DependenciaForm(instance=dependencia)

    return render(request, 'almacen/crear_dependencia.html', {'form': form, 'dependencias': dependencias})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def crear_ubicacion(request):
    ubicaciones = Ubicacion.objects.all()
    form = UbicacionForm(request.POST or None)

    if form.is_valid():
        # Guardamos el formulario sin el commit para poder manipular el objeto antes de guardarlo
        ubicacion = form.save(commit=False)
        
        # Solo establecer 'activo' como True si es una nueva ubicación (sin pk)
        if not ubicacion.pk:
            ubicacion.activo = True
        
        # Guardamos la ubicación
        ubicacion.save()
        
        # Redirigimos a la misma página para mostrar la nueva ubicación creada
        return redirect('almacen:crear_ubicacion')

    return render(request, 'almacen/crear_ubicacion.html', {'form': form, 'ubicaciones': ubicaciones})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def editar_ubicacion(request, pk):
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    form = UbicacionForm(request.POST or None, instance=ubicacion)
    if form.is_valid():
        form.save()
        return redirect('almacen:crear_ubicacion')  # Redirige a la vista de creación
    return render(request, 'almacen/editar_ubicacion.html', {'form': form, 'ubicaciones': Ubicacion.objects.all()})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def buscar_proveedor_por_nit(request, nit):
    try:
        proveedor = Proveedor.objects.get(nit=nit)
        # Devuelve los datos del proveedor en formato JSON
        return JsonResponse({
            'nombre': proveedor.nombre,
            'telefono': proveedor.telefono,
            'direccion': proveedor.direccion,
        })
    except Proveedor.DoesNotExist:
        # Si no se encuentra el proveedor, devuelve un error
        return JsonResponse({'error': 'Proveedor no encontrado'}, status=404)


def eliminar_detalle_factura(request, detalle_id):
    if request.method == "POST":
        detalle = get_object_or_404(DetalleFactura, id=detalle_id)
        detalle.delete()
        return JsonResponse({'success': True})  # Respuesta JSON para indicar éxito
    return JsonResponse({'success': False}, status=400)  # Respuesta en caso de error

def obtener_detalle_factura(request, detalle_id):
    detalle = get_object_or_404(DetalleFactura, id=detalle_id)
    data = {
        'articulo': detalle.articulo.id,
        'cantidad': detalle.cantidad,
        'precio_unitario': detalle.precio_unitario,
        'renglon': detalle.renglon,
        'fecha_vencimiento': detalle.fecha_vencimiento.strftime('%Y-%m-%d') if detalle.fecha_vencimiento else '',
    }
    return JsonResponse(data)

@login_required
@grupo_requerido('Administrador', 'Almacen')  
def editar_detalle_factura(request):
    if request.method == 'POST':
        detalle_id = request.POST.get("detalle_id")  # Obtener el ID del detalle
        detalle = get_object_or_404(DetalleFactura, id=detalle_id)  # Buscar el detalle existente

        # Obtener el artículo y validar si requiere fecha de vencimiento
        articulo_id = request.POST.get("articulo")
        articulo = get_object_or_404(Articulo, id=articulo_id)

        # Verificar si el artículo requiere fecha de vencimiento
        if articulo.requiere_vencimiento:
            fecha_vencimiento = request.POST.get("fecha_vencimiento")
            if not fecha_vencimiento:  # Si no se envió la fecha de vencimiento
                raise ValidationError("Este artículo requiere una fecha de vencimiento.")
        else:
            fecha_vencimiento = None  # No es necesario asignar fecha si no se requiere

        # Actualizar los campos del detalle
        detalle.articulo_id = articulo_id
        detalle.cantidad = int(request.POST.get("cantidad"))  # Convertir a entero
        detalle.precio_unitario = float(request.POST.get("precio_unitario"))  # Convertir a flotante
        detalle.renglon = request.POST.get("renglon")
        detalle.precio_total = detalle.cantidad * detalle.precio_unitario  # Multiplicación correcta
        
        # Si se requiere fecha de vencimiento y está presente, guardarla
        if fecha_vencimiento:
            detalle.fecha_vencimiento = fecha_vencimiento
        
        detalle.save()

        # Redirigir después de la actualización
        return redirect('almacen:agregar_detalle_factura', form1h_id=detalle.form1h.id)

    return HttpResponseNotAllowed(['POST'])

@login_required
@grupo_requerido('Administrador', 'Almacen')
def user_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('new_password')
            user.set_password(password)
            user.save()

            group = form.cleaned_data.get('group')
            user.groups.add(group)

            # ✅ Espera a que la señal cree el perfil automáticamente
            foto = form.cleaned_data.get('foto')
            try:
                perfil = user.perfil  # accede al perfil creado por la señal
                if foto:
                    perfil.foto = foto
                    perfil.save()
            except Perfil.DoesNotExist:
                # Fallback solo si la señal falló (raro)
                Perfil.objects.create(user=user, foto=foto)

            messages.success(request, 'Usuario creado correctamente.')
            return redirect('almacen:user_create')
    else:
        form = UserCreateForm()

    users = User.objects.all()
    return render(request, 'almacen/user_form_create.html', {'form': form, 'users': users})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def user_edit(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    try:
        perfil = user.perfil
    except Perfil.DoesNotExist:
        perfil = Perfil(user=user)

    if request.method == 'POST':
        form_user = UserEditForm(request.POST, instance=user)
        form_perfil = PerfilForm(request.POST, request.FILES, instance=perfil)
        if form_user.is_valid() and form_perfil.is_valid():
            user = form_user.save(commit=False)
            user.save()

            # Actualizar grupo: limpiar y agregar el nuevo grupo
            group = form_user.cleaned_data.get('group')
            if group:
                user.groups.clear()
                user.groups.add(group)

            perfil = form_perfil.save(commit=False)
            perfil.user = user
            perfil.save()

            messages.success(request, 'Usuario editado correctamente.')
            return redirect('almacen:user_create')
    else:
        form_user = UserEditForm(instance=user)
        form_perfil = PerfilForm(instance=perfil)

    context = {
        'form': form_user,
        'perfil_form': form_perfil,
        'users': User.objects.all(),
    }
    return render(request, 'almacen/user_form_edit.html', context)



@login_required
@grupo_requerido('Administrador', 'Almacen')
def perfil_edit(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    try:
        perfil = user.perfil
    except Perfil.DoesNotExist:
        perfil = Perfil(user=user)
    
    if request.method == 'POST':
        form = PerfilForm(request.POST, request.FILES, instance=perfil)
        if form.is_valid():
            form.save()
            return redirect('almacen:user_edit', user_id=user.id)
    else:
        form = PerfilForm(instance=perfil)
    
    return render(request, 'almacen/perfil_edit.html', {'form': form, 'user': user})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        return redirect('almacen:user_create')  # Redirige a la misma página para mostrar la lista actualizada
    return render(request, 'almacen/user_confirm_delete.html', {'user': user})


def home(request):
    return render(request, 'almacen/login.html')

from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Q, Sum
import json

@login_required
@grupo_requerido('Administrador', 'Almacen')
def dahsboard(request):
    # Totales de artículos activos/inactivos
    totales_articulos = Articulo.objects.aggregate(
        activos=Count('id', filter=Q(activo=True)),
        inactivos=Count('id', filter=Q(activo=False))
    )

    # Suma de Kardex: entradas/salidas
    ingresos = Kardex.objects.filter(tipo_movimiento='INGRESO').aggregate(total=Sum('cantidad'))['total'] or 0
    salidas = Kardex.objects.filter(tipo_movimiento='SALIDA').aggregate(total=Sum('cantidad'))['total'] or 0

    # Requerimientos por estado
    estados_orden = ['pendiente', 'despachado', 'rechazado', 'enviado']
    req_por_estado = Requerimiento.objects.values('estado').annotate(cuenta=Count('id'))
    req_por_estado = sorted(
        list(req_por_estado),
        key=lambda r: estados_orden.index(r['estado']) if r['estado'] in estados_orden else len(estados_orden)
    )

    labels_req = [r['estado'].capitalize() for r in req_por_estado]
    data_req = [r['cuenta'] for r in req_por_estado]
    total_reqs = sum(data_req)

    # Evolución semanal de requerimientos (este mes)
    today = timezone.now().date()
    stm = today.replace(day=1)
    etm = (stm + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    req_sem = Requerimiento.objects.annotate(
        week=TruncWeek('fecha_creacion')
    ).filter(fecha_creacion__range=(stm, etm)).values('week').annotate(cuenta=Count('id')).order_by('week')

    semanas = [r['week'].strftime('%b %d') for r in req_sem]
    cantidad_sem = [r['cuenta'] for r in req_sem]

    # NUEVO: Artículos por vencer (en los próximos 30 días)
    limite = today + timedelta(days=30)

    detalles_por_vencer = DetalleFactura.objects.filter(
        articulo__requiere_vencimiento=True,
        fecha_vencimiento__gte=today,
        fecha_vencimiento__lte=limite,
    ).values('fecha_vencimiento').annotate(cantidad=Count('id')).order_by('fecha_vencimiento')

    fechas_vencimiento = [item['fecha_vencimiento'].strftime('%Y-%m-%d') for item in detalles_por_vencer]
    cantidades_vencimiento = [item['cantidad'] for item in detalles_por_vencer]
    total_por_vencer = sum(cantidades_vencimiento)

    context = {
        'totales_articulos': totales_articulos,
        'ingresos': ingresos,
        'salidas': salidas,
        'labels_req': json.dumps(labels_req),
        'data_req': json.dumps(data_req),
        'semanas': json.dumps(semanas),
        'cantidad_sem': json.dumps(cantidad_sem),
        'total_por_vencer': total_por_vencer,
        'fechas_vencimiento': json.dumps(fechas_vencimiento),
        'cantidades_vencimiento': json.dumps(cantidades_vencimiento),
    }

    return render(request, 'almacen/dashboard.html', context)

@login_required
@grupo_requerido('Administrador', 'Almacen')
def articulos_por_vencer(request):
    today = timezone.now().date()
    limite = today + timedelta(days=30)

    # Obtener todas las salidas agrupadas por DetalleFactura (fuente_factura)
    salidas_dict = Kardex.objects.filter(
        tipo_movimiento='SALIDA',
        fuente_factura__isnull=False
    ).values('fuente_factura').annotate(
        total=Coalesce(Sum('cantidad'), 0)
    )

    # Convertir a diccionario para acceso rápido
    salidas_map = {item['fuente_factura']: item['total'] for item in salidas_dict}

    # Artículos por vencer
    articulos_vencer = DetalleFactura.objects.filter(
        articulo__requiere_vencimiento=True,
        fecha_vencimiento__gte=today,
        fecha_vencimiento__lte=limite,
        form1h__estado='confirmado'
    ).select_related('articulo', 'form1h').order_by('fecha_vencimiento')

    # Artículos vencidos
    articulos_vencidos = DetalleFactura.objects.filter(
        articulo__requiere_vencimiento=True,
        fecha_vencimiento__lt=today,
        form1h__estado='confirmado'
    ).select_related('articulo', 'form1h').order_by('-fecha_vencimiento')

    # Calcular el stock restante manualmente (sin Subquery)
    for detalle in list(articulos_vencer) + list(articulos_vencidos):
        total_salidas = salidas_map.get(detalle.pk, 0)
        detalle.stock_restante = detalle.cantidad - total_salidas

    context = {
        'articulos_vencer': articulos_vencer,
        'articulos_vencidos': articulos_vencidos,
        'today': today,
        'limite': limite,
    }

    return render(request, 'almacen/articulos_por_vencer.html', context)




def signout(request):
    logout(request)
    return redirect('almacen:signin')


def signin(request):  
    institucion = Institucion.objects.first()
    if request.method == 'GET':
        # Deberías instanciar el AuthenticationForm correctamente
        return render(request, 'almacen/login.html', {
            'form': AuthenticationForm(),
            'institucion': institucion,
        })
    else:
        # Se instancia AuthenticationForm con los datos del POST para mantener el estado
        form = AuthenticationForm(request, data=request.POST)
        
        if form.is_valid():
            # El método authenticate devuelve el usuario si es válido
            user = form.get_user()
            
            # Si el usuario es encontrado, se inicia sesión
            auth_login(request, user)
            
            # Ahora verificamos los grupos
            for g in user.groups.all():
                print(g.name)
                if g.name == 'Administrador':
                    return redirect('almacen:dahsboard')
                elif g.name == 'Departamento':
                    return redirect('almacen:crear_requerimiento')
                elif g.name == 'Gestor':
                    return redirect('almacen:listado_solicitudes_gestor')
            # Si no se encuentra el grupo adecuado, se redirige a una página por defecto
            return redirect('dahsboard')
        else:
            # Si el formulario no es válido, se retorna con el error
            return render(request, 'almacen/login.html', {
                'form': form,  # Pasamos el formulario con los errores
                'error': 'Usuario o contraseña incorrectos',
                'institucion': institucion,
            })


def _parse_fecha_libro_mayor(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return None


def _moneda(valor):
    return Decimal(valor or 0).quantize(Decimal('0.01'))


def _normalizar_rango_renglones(inicio=None, fin=None):
    try:
        inicio = int(inicio)
    except (TypeError, ValueError):
        inicio = 300
    try:
        fin = int(fin)
    except (TypeError, ValueError):
        fin = 399
    inicio = max(300, min(inicio, 399))
    fin = max(300, min(fin, 399))
    if inicio > fin:
        inicio, fin = fin, inicio
    return inicio, fin


def _departamentos_permitidos_libro_mayor(usuario):
    if usuario.is_superuser or usuario.groups.filter(name__in=['Administrador', 'Almacen']).exists():
        return None
    if usuario.groups.filter(name__in=['Departamento', 'Gestor']).exists():
        return list(UsuarioDepartamento.objects.filter(usuario=usuario).values_list('departamento_id', flat=True))
    return []


def _movimiento_corresponde_departamento(movimiento, departamento_ids):
    if departamento_ids is None:
        return True
    if not departamento_ids:
        return False
    if movimiento.fuente_asignacion_id:
        return movimiento.fuente_asignacion.destino_id in departamento_ids
    if movimiento.fuente_despacho_id and movimiento.fuente_despacho.requerimiento_id:
        return movimiento.fuente_despacho.requerimiento.departamento_id in departamento_ids
    return movimiento.tipo_movimiento == 'INGRESO'


def _costo_entrada_kardex(movimiento):
    if movimiento.fuente_factura_id and movimiento.fuente_factura:
        precio_unitario = _moneda(movimiento.fuente_factura.precio_unitario)
        precio_total = _moneda(movimiento.fuente_factura.precio_total)
        sin_costo = precio_unitario <= 0
        if precio_total > 0 and movimiento.fuente_factura.cantidad == movimiento.cantidad:
            return precio_unitario, precio_total, sin_costo
        return precio_unitario, _moneda(precio_unitario * Decimal(movimiento.cantidad or 0)), sin_costo
    return Decimal('0.00'), Decimal('0.00'), True


def _datos_departamento_movimiento(movimiento):
    if movimiento.fuente_asignacion_id:
        return movimiento.fuente_asignacion.destino.nombre
    if movimiento.fuente_despacho_id and movimiento.fuente_despacho.requerimiento_id:
        return movimiento.fuente_despacho.requerimiento.departamento.nombre
    return ''


def _calcular_movimientos_valorizados(articulo, departamento_ids, fecha_inicio=None, fecha_fin=None):
    saldo_cantidad = 0
    saldo_monto = Decimal('0.00')
    resumen = {
        'saldo_inicial_cantidad': 0,
        'saldo_inicial_monto': Decimal('0.00'),
        'entradas_cantidad': 0,
        'entradas_monto': Decimal('0.00'),
        'salidas_cantidad': 0,
        'salidas_monto': Decimal('0.00'),
        'sin_costo': False,
        'movimientos': [],
        'tuvo_movimientos': False,
    }

    movimientos = Kardex.objects.filter(articulo=articulo).select_related(
        'fuente_factura', 'fuente_asignacion__destino', 'fuente_despacho__requerimiento__departamento'
    ).order_by('fecha', 'id')

    for mov in movimientos:
        if not _movimiento_corresponde_departamento(mov, departamento_ids):
            continue
        resumen['tuvo_movimientos'] = True
        cantidad = mov.cantidad or 0
        cantidad_decimal = Decimal(cantidad)
        es_ingreso = mov.tipo_movimiento in ['INGRESO', 'ENTRADA']
        es_salida = mov.tipo_movimiento in ['SALIDA', 'EGRESO']
        costo_unitario = Decimal('0.00')
        monto = Decimal('0.00')
        sin_costo_movimiento = False

        if es_ingreso:
            costo_unitario, monto, sin_costo_movimiento = _costo_entrada_kardex(mov)
            saldo_cantidad += cantidad
            saldo_monto = _moneda(saldo_monto + monto)
        elif es_salida:
            costo_unitario = _moneda(saldo_monto / Decimal(saldo_cantidad)) if saldo_cantidad > 0 else Decimal('0.00')
            monto = _moneda(costo_unitario * cantidad_decimal)
            if costo_unitario == 0 and cantidad > 0:
                sin_costo_movimiento = True
            saldo_cantidad -= cantidad
            saldo_monto = _moneda(saldo_monto - monto)
        else:
            continue

        if sin_costo_movimiento:
            resumen['sin_costo'] = True

        fecha_mov = mov.fecha.date()
        dentro_rango = (not fecha_inicio or fecha_mov >= fecha_inicio) and (not fecha_fin or fecha_mov <= fecha_fin)
        if fecha_inicio and fecha_mov < fecha_inicio:
            resumen['saldo_inicial_cantidad'] = saldo_cantidad
            resumen['saldo_inicial_monto'] = saldo_monto
            continue
        if not dentro_rango:
            continue

        if es_ingreso:
            resumen['entradas_cantidad'] += cantidad
            resumen['entradas_monto'] = _moneda(resumen['entradas_monto'] + monto)
        elif es_salida:
            resumen['salidas_cantidad'] += cantidad
            resumen['salidas_monto'] = _moneda(resumen['salidas_monto'] + monto)

        resumen['movimientos'].append({
            'fecha': mov.fecha,
            'tipo_movimiento': mov.tipo_movimiento,
            'referencia': mov.numero_kardex,
            'departamento': _datos_departamento_movimiento(mov),
            'entrada_cantidad': cantidad if es_ingreso else 0,
            'entrada_monto': monto if es_ingreso else Decimal('0.00'),
            'salida_cantidad': cantidad if es_salida else 0,
            'salida_monto': monto if es_salida else Decimal('0.00'),
            'saldo_cantidad': saldo_cantidad,
            'saldo_monto': saldo_monto,
            'costo_unitario': costo_unitario,
            'sin_costo': sin_costo_movimiento,
            'revisar': saldo_cantidad < 0 or saldo_monto < 0,
            'observacion': mov.observacion,
        })

    if not fecha_inicio:
        resumen['saldo_inicial_cantidad'] = 0
        resumen['saldo_inicial_monto'] = Decimal('0.00')
    resumen['saldo_final_cantidad'] = resumen['saldo_inicial_cantidad'] + resumen['entradas_cantidad'] - resumen['salidas_cantidad']
    resumen['saldo_final_monto'] = _moneda(resumen['saldo_inicial_monto'] + resumen['entradas_monto'] - resumen['salidas_monto'])
    resumen['costo_promedio'] = _moneda(resumen['saldo_final_monto'] / Decimal(resumen['saldo_final_cantidad'])) if resumen['saldo_final_cantidad'] > 0 else Decimal('0.00')
    resumen['revisar'] = resumen['saldo_final_cantidad'] < 0 or resumen['saldo_final_monto'] < 0
    return resumen


def generar_datos_libro_mayor(fecha_inicio=None, fecha_fin=None, renglon_inicio=300, renglon_fin=399, articulo=None, departamento=None, usuario=None, incluir_sin_movimientos=False):
    renglon_inicio, renglon_fin = _normalizar_rango_renglones(renglon_inicio, renglon_fin)
    fecha_inicio = _parse_fecha_libro_mayor(fecha_inicio) if isinstance(fecha_inicio, str) else fecha_inicio
    fecha_fin = _parse_fecha_libro_mayor(fecha_fin) if isinstance(fecha_fin, str) else fecha_fin

    departamento_ids = None
    if usuario is not None:
        departamento_ids = _departamentos_permitidos_libro_mayor(usuario)
    if departamento:
        try:
            departamento_id = int(departamento)
            departamento_ids = [departamento_id] if departamento_ids is None or departamento_id in departamento_ids else []
        except (TypeError, ValueError):
            departamento_ids = []

    articulos = Articulo.objects.select_related('unidad_medida').order_by('renglon_presupuestario', 'nombre')
    if articulo:
        articulos = articulos.filter(id=articulo)

    filas = []
    for art in articulos:
        renglon_texto = (art.renglon_presupuestario or '').strip()
        if not renglon_texto.isdigit():
            continue
        renglon_numero = int(renglon_texto)
        if not (renglon_inicio <= renglon_numero <= renglon_fin):
            continue
        calculo = _calcular_movimientos_valorizados(art, departamento_ids, fecha_inicio, fecha_fin)
        if not calculo['tuvo_movimientos'] and not incluir_sin_movimientos:
            continue
        filas.append({
            'renglon': str(renglon_numero),
            'codigo': art.codigo,
            'articulo': art.nombre,
            'unidad': str(art.unidad_medida) if art.unidad_medida else '',
            'saldo_inicial_cantidad': calculo['saldo_inicial_cantidad'],
            'saldo_inicial_monto': calculo['saldo_inicial_monto'],
            'entradas_cantidad': calculo['entradas_cantidad'],
            'entradas_monto': calculo['entradas_monto'],
            'salidas_cantidad': calculo['salidas_cantidad'],
            'salidas_monto': calculo['salidas_monto'],
            'saldo_final_cantidad': calculo['saldo_final_cantidad'],
            'saldo_final_monto': calculo['saldo_final_monto'],
            'costo_promedio': calculo['costo_promedio'],
            'sin_costo': calculo['sin_costo'],
            'revisar': calculo['revisar'],
            'articulo_id': art.id,
        })

    totales = {
        'total_renglones': len({fila['renglon'] for fila in filas}),
        'total_articulos': len(filas),
        'total_entradas_cantidad': sum(fila['entradas_cantidad'] for fila in filas),
        'total_entradas_monto': _moneda(sum((fila['entradas_monto'] for fila in filas), Decimal('0.00'))),
        'total_salidas_cantidad': sum(fila['salidas_cantidad'] for fila in filas),
        'total_salidas_monto': _moneda(sum((fila['salidas_monto'] for fila in filas), Decimal('0.00'))),
        'saldo_final_cantidad': sum(fila['saldo_final_cantidad'] for fila in filas),
        'saldo_final_monto': _moneda(sum((fila['saldo_final_monto'] for fila in filas), Decimal('0.00'))),
    }
    filtros = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'renglon_inicio': renglon_inicio,
        'renglon_fin': renglon_fin,
        'articulo': str(articulo or ''),
        'departamento': str(departamento or ''),
        'incluir_sin_movimientos': incluir_sin_movimientos,
    }
    return {'filas': filas, 'totales': totales, 'filtros': filtros}



@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento', 'Gestor')
def libro_mayor(request):
    datos = generar_datos_libro_mayor(
        fecha_inicio=request.GET.get('fecha_inicio'),
        fecha_fin=request.GET.get('fecha_fin'),
        renglon_inicio=request.GET.get('renglon_inicio', 300),
        renglon_fin=request.GET.get('renglon_fin', 399),
        articulo=request.GET.get('articulo'),
        departamento=request.GET.get('departamento'),
        usuario=request.user,
        incluir_sin_movimientos=request.GET.get('incluir_sin_movimientos') == 'on',
    )
    departamentos_ids = _departamentos_permitidos_libro_mayor(request.user)
    departamentos = Departamento.objects.all() if departamentos_ids is None else Departamento.objects.filter(id__in=departamentos_ids)
    return render(request, 'almacen/libro_mayor.html', {
        **datos,
        'articulos': Articulo.objects.filter(renglon_presupuestario__regex=r'^3[0-9][0-9]$').order_by('codigo', 'nombre'),
        'departamentos': departamentos,
    })


@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento', 'Gestor')
def libro_mayor_pdf(request):
    datos = generar_datos_libro_mayor(request.GET.get('fecha_inicio'), request.GET.get('fecha_fin'), request.GET.get('renglon_inicio', 300), request.GET.get('renglon_fin', 399), request.GET.get('articulo'), request.GET.get('departamento'), request.user, request.GET.get('incluir_sin_movimientos') == 'on')
    html_string = render_to_string('almacen/libro_mayor_pdf.html', {**datos, 'institucion': Institucion.objects.first(), 'generado': timezone.now()})
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="libro_mayor.pdf"'
    return response


@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento', 'Gestor')
def libro_mayor_excel(request):
    datos = generar_datos_libro_mayor(request.GET.get('fecha_inicio'), request.GET.get('fecha_fin'), request.GET.get('renglon_inicio', 300), request.GET.get('renglon_fin', 399), request.GET.get('articulo'), request.GET.get('departamento'), request.user, request.GET.get('incluir_sin_movimientos') == 'on')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Libro Mayor'
    ws.append(['Libro Mayor'])
    ws.append(['Generado', timezone.now().strftime('%d/%m/%Y %H:%M')])
    ws.append(['Rango renglones', f"{datos['filtros']['renglon_inicio']} - {datos['filtros']['renglon_fin']}"])
    ws.append(['Fecha inicio', datos['filtros']['fecha_inicio'] or ''])
    ws.append(['Fecha fin', datos['filtros']['fecha_fin'] or ''])
    ws.append([])
    headers = [
        'Renglón', 'Código', 'Artículo', 'Unidad',
        'Saldo Inicial Cantidad', 'Saldo Inicial Monto',
        'Entradas Cantidad', 'Entradas Monto',
        'Salidas Cantidad', 'Salidas Monto',
        'Saldo Final Cantidad', 'Saldo Final Monto', 'Costo Promedio'
    ]
    ws.append(headers)
    for cell in ws[7]:
        cell.font = Font(bold=True)
    for fila in datos['filas']:
        ws.append([
            fila['renglon'], fila['codigo'], fila['articulo'], fila['unidad'],
            fila['saldo_inicial_cantidad'], fila['saldo_inicial_monto'],
            fila['entradas_cantidad'], fila['entradas_monto'],
            fila['salidas_cantidad'], fila['salidas_monto'],
            fila['saldo_final_cantidad'], fila['saldo_final_monto'], fila['costo_promedio']
        ])
    ws.append([])
    ws.append([
        'TOTAL', '', '', '', '', '',
        datos['totales']['total_entradas_cantidad'], datos['totales']['total_entradas_monto'],
        datos['totales']['total_salidas_cantidad'], datos['totales']['total_salidas_monto'],
        datos['totales']['saldo_final_cantidad'], datos['totales']['saldo_final_monto'], ''
    ])
    for row in ws.iter_rows(min_row=8, min_col=6, max_col=13):
        for cell in row:
            if cell.column in [6, 8, 10, 12, 13]:
                cell.number_format = '"Q"#,##0.00'
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(len(str(cell.value or '')) for cell in col) + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="libro_mayor.xlsx"'
    wb.save(response)
    return response


@login_required
@grupo_requerido('Administrador', 'Almacen', 'Departamento', 'Gestor')
def libro_mayor_movimientos(request, articulo_id):
    articulo = get_object_or_404(Articulo, id=articulo_id)
    fecha_inicio = _parse_fecha_libro_mayor(request.GET.get('fecha_inicio'))
    fecha_fin = _parse_fecha_libro_mayor(request.GET.get('fecha_fin'))
    departamento_ids = _departamentos_permitidos_libro_mayor(request.user)
    if request.GET.get('departamento'):
        try:
            departamento_id = int(request.GET.get('departamento'))
            departamento_ids = [departamento_id] if departamento_ids is None or departamento_id in departamento_ids else []
        except (TypeError, ValueError):
            departamento_ids = []
    calculo = _calcular_movimientos_valorizados(articulo, departamento_ids, fecha_inicio, fecha_fin)
    return render(request, 'almacen/libro_mayor_movimientos.html', {
        'articulo': articulo,
        'movimientos': calculo['movimientos'],
        'filtros': {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin},
    })

@login_required
@grupo_requerido('Administrador', 'Almacen')
def articulo_disponibilidad_json(request, articulo_id):
    articulo = get_object_or_404(Articulo, pk=articulo_id)
    existencia_total = Decimal(existencia_general_articulo(articulo))
    asignado_divisiones = DivisionArticulo.objects.filter(articulo=articulo, activo=True).aggregate(total=Sum('cantidad_asignada'))['total'] or Decimal('0')
    asignado_ubicaciones = DivisionArticuloUbicacion.objects.filter(division_articulo__articulo=articulo, activo=True).aggregate(total=Sum('cantidad_asignada'))['total'] or Decimal('0')
    reservado = DivisionArticuloUbicacion.objects.filter(division_articulo__articulo=articulo, activo=True).aggregate(total=Sum('cantidad_reservada'))['total'] or Decimal('0')
    disponible_sin_reservar = max(Decimal('0'), existencia_total - reservado)
    disponible_para_division = max(Decimal('0'), existencia_total - asignado_divisiones)
    return JsonResponse({
        'codigo': articulo.codigo,
        'nombre': articulo.nombre,
        'existencia_total': str(existencia_total),
        'asignado_divisiones': str(asignado_divisiones),
        'asignado_ubicaciones': str(asignado_ubicaciones),
        'reservado': str(reservado),
        'disponible_sin_reservar': str(disponible_sin_reservar),
        'disponible_para_division': str(disponible_para_division),
        'revisar': disponible_sin_reservar <= 0,
    })

@login_required
@grupo_requerido('Administrador', 'Almacen')
@require_POST
def desasignar_articulo_ubicacion(request, asignacion_id):
    asignacion = get_object_or_404(DivisionArticuloUbicacion, pk=asignacion_id)
    redirect_to = request.POST.get('next') or reverse('almacen:division_detail', kwargs={'pk': asignacion.division_articulo.division_id})
    try:
        cantidad = Decimal(request.POST.get('cantidad_desasignar') or '0')
    except Exception:
        messages.error(request, 'Cantidad inválida para desasignar.')
        return redirect(redirect_to)
    disponible_para_desasignar = max(Decimal('0'), asignacion.cantidad_asignada - asignacion.cantidad_reservada - asignacion.cantidad_consumida)
    if cantidad <= 0:
        messages.error(request, 'La cantidad a desasignar debe ser mayor a cero.')
        return redirect(redirect_to)
    if cantidad > disponible_para_desasignar:
        messages.error(request, 'No puede desasignar más de la cantidad disponible sin reservar.')
        return redirect(redirect_to)
    asignacion.cantidad_asignada -= cantidad
    if asignacion.cantidad_asignada <= 0 and asignacion.cantidad_reservada <= 0 and asignacion.cantidad_consumida <= 0:
        asignacion.activo = False
    asignacion.save(update_fields=['cantidad_asignada', 'activo', 'fecha_actualizacion'])
    messages.success(request, f'Se desasignaron {cantidad} unidades correctamente.')
    return redirect(redirect_to)

@login_required
@grupo_requerido('Administrador', 'Almacen')
def division_list(request):
    divisiones = DivisionAlmacen.objects.order_by('nombre')
    for division in divisiones:
        division.total_ubicaciones_listado = DivisionUbicacion.objects.filter(
            division=division,
            activa=True
        ).values('ubicacion_id').distinct().count()
        division.total_articulos_listado = DivisionArticulo.objects.filter(
            division=division,
            activo=True
        ).values('articulo_id').distinct().count()
        division.total_reservado_listado = DivisionArticuloUbicacion.objects.filter(
            division_articulo__division=division,
            division_articulo__activo=True,
            activo=True
        ).aggregate(
            total=Coalesce(Sum('cantidad_reservada'), Decimal('0.00'))
        )['total']
    return render(request, 'almacen/divisiones/list.html', {'divisiones': divisiones})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def division_create(request):
    form = DivisionAlmacenForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        division = form.save(commit=False); division.creado_por = request.user; division.save()
        messages.success(request, 'División creada correctamente.')
        return redirect('almacen:division_detail', pk=division.pk)
    return render(request, 'almacen/divisiones/form.html', {'form': form, 'titulo': 'Crear División'})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def division_update(request, pk):
    division = get_object_or_404(DivisionAlmacen, pk=pk)
    form = DivisionAlmacenForm(request.POST or None, instance=division)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'División actualizada correctamente.')
        return redirect('almacen:division_detail', pk=division.pk)
    return render(request, 'almacen/divisiones/form.html', {'form': form, 'titulo': 'Editar División'})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def division_detail(request, pk):
    division = get_object_or_404(DivisionAlmacen, pk=pk)
    ubicacion_form = DivisionUbicacionForm(prefix='ubicacion')
    articulo_form = DivisionArticuloForm(prefix='articulo')
    if request.method == 'POST':
        accion = request.POST.get('accion')
        try:
            if accion == 'agregar_ubicacion':
                ubicacion_form = DivisionUbicacionForm(request.POST, prefix='ubicacion')
                if ubicacion_form.is_valid():
                    obj = ubicacion_form.save(commit=False); obj.division = division; obj.save(); messages.success(request, 'Ubicación agregada.')
            elif accion == 'agregar_articulo':
                articulo_form = DivisionArticuloForm(request.POST, prefix='articulo')
                if articulo_form.is_valid():
                    obj = articulo_form.save(commit=False); obj.division = division; obj.asignado_por = request.user; obj.save(); messages.success(request, 'Artículo agregado.')
            elif accion == 'desactivar_ubicacion':
                DivisionUbicacion.objects.filter(pk=request.POST.get('id'), division=division).update(activa=False); messages.success(request, 'Ubicación desactivada.')
            elif accion == 'desactivar_articulo':
                DivisionArticulo.objects.filter(pk=request.POST.get('id'), division=division).update(activo=False); messages.success(request, 'Artículo desactivado.')
            elif accion == 'desasignar_ubicacion':
                messages.error(request, 'Use la opción de desasignación parcial e indique una cantidad.')
        except ValidationError as e:
            messages.error(request, e.message_dict if hasattr(e, 'message_dict') else e.messages)
        return redirect('almacen:division_detail', pk=division.pk)
    return render(request, 'almacen/divisiones/detail.html', {'division': division, 'ubicacion_form': ubicacion_form, 'articulo_form': articulo_form})

@login_required
@grupo_requerido('Administrador', 'Almacen')
def division_control(request):
    articulos = DivisionArticulo.objects.select_related('division','articulo','articulo__categoria').filter(activo=True).order_by('division__nombre','articulo__nombre')
    asignaciones = DivisionArticuloUbicacion.objects.filter(activo=True).select_related('division_articulo__division', 'division_articulo__articulo', 'ubicacion').order_by('division_articulo__division__nombre', 'ubicacion__nombre', 'division_articulo__articulo__nombre')
    return render(request, 'almacen/divisiones/control.html', {'articulos': articulos, 'asignaciones': asignaciones})

@login_required
@grupo_requerido('Gestor', 'Departamento')
def mis_divisiones(request):
    deptos = UsuarioDepartamento.objects.filter(usuario=request.user).values_list('departamento_id', flat=True)
    articulos = DivisionArticulo.objects.filter(activo=True, division__ubicaciones__ubicacion_id__in=deptos, division__ubicaciones__activa=True).select_related('division','articulo','articulo__categoria','articulo__unidad_medida').distinct()
    asignaciones = DivisionArticuloUbicacion.objects.filter(ubicacion_id__in=deptos, activo=True)
    asignacion_map = {(a.division_articulo_id, a.ubicacion_id): a for a in asignaciones}
    return render(request, 'almacen/divisiones/mis_divisiones.html', {'articulos': articulos, 'deptos': deptos, 'asignacion_map': asignacion_map})

@login_required
@grupo_requerido('Gestor', 'Departamento')
def autoasignar_division_articulo(request, pk):
    da = get_object_or_404(DivisionArticulo, pk=pk, activo=True)
    depto = UsuarioDepartamento.objects.filter(usuario=request.user, departamento__divisiones_almacen__division=da.division, departamento__divisiones_almacen__activa=True).first()
    if not depto:
        messages.error(request, 'No pertenece a esta división.'); return redirect('almacen:mis_divisiones')
    if request.method == 'POST':
        cantidad = Decimal(request.POST.get('cantidad') or '0')
        obj = DivisionArticuloUbicacion.objects.filter(division_articulo=da, ubicacion=depto.departamento).first()
        if obj:
            obj.cantidad_asignada += cantidad
        else:
            obj = DivisionArticuloUbicacion(division_articulo=da, ubicacion=depto.departamento, cantidad_asignada=cantidad)
        obj.activo = True; obj.asignado_por = request.user
        try:
            obj.save(); messages.success(request, 'Artículo autoasignado correctamente.')
        except ValidationError as e:
            messages.error(request, e.message_dict if hasattr(e, 'message_dict') else e.messages)
    return redirect('almacen:mis_divisiones')

@login_required
@grupo_requerido('Gestor', 'Departamento')
def mis_articulos_asignados(request):
    deptos = UsuarioDepartamento.objects.filter(usuario=request.user).values_list('departamento_id', flat=True)
    asignaciones = DivisionArticuloUbicacion.objects.filter(ubicacion_id__in=deptos, activo=True).select_related('division_articulo__division','division_articulo__articulo')
    return render(request, 'almacen/divisiones/mis_articulos_asignados.html', {'asignaciones': asignaciones})
